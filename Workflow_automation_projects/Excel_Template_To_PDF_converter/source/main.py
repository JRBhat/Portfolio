

import os
import subprocess
from time import sleep
import sys
import re

from openpyxl import load_workbook, utils
from ImageAnalysis import Util

from Internal_Imports_Stable import InternalImport
import Common_Functions_Stable as CFS 

from Create_Excel_template import create_excel 
from Randomization_Template_Stable import randomize_std, randomize_custom_rand_alllightingcode
from Randomization_Template_Transposed_Stable import randomize_transp
from Column_Name_Per_Row_Template import give_each_row_columnames

from Insert_Description_Stable import Insert_description_file
from File_renamers.Visia_file_rename_before_PDF_creation import rename_visia_files
from Templates import Standard, Transpose, Custom

MODIFY_TEMPLATE_SHEET = "Modify_Template_here"
CUSTOM_TEMPLATE_FLAG = "*"
COLUMN_NAMES_FLAG = "§"
COLUMN_SCAN_LIMIT = 20

def main():
    """
    the heart of the program containing bulk of the logic 
    """
    Imp = InternalImport()
    
    Imp.validated_path, _ = CFS.convert_and_scale_to_standard_jpgs(Imp.file_extension, 
                                                                                Imp.path_for_validation)
    
    subprocess.Popen(f"d: && start {Imp.validated_path}", shell=True)
    input("Conversion complete. Please check the created images in the opened folder and then press enter")


    checker = input("Are the images transformed correctly...(y/n)?")

    while True:
        if checker.lower() == "y":
            break
        elif checker.lower() == "n": 
            print("Something wrong with the images. Stopping program..")
            sys.exit(1)
        checker = input("Are the images transformed correctly...(y/n)?")
        
    filenamelist = Util.getAllFiles(Imp.validated_path, "*.jpg", depth=-1)

    # ............... check for visia files.........
    is_visia = False
    filename_regex = re.compile(rf'{Imp.filename_mask}')
    visia_regex = re.compile(r"([0-9]*)_([A-Za-z0-9 ]*)_([a-zA-Z \-]*)_([a-zA-Z0-9 \-]*)")

    first_img_filename = [fn for fn in filenamelist if fn.endswith(".jpg") or fn.endswith(".JPG")][0].split("\\")[-1]

    if not re.match(filename_regex, first_img_filename):
        if visia_regex.search(first_img_filename).group(0):
            is_visia = True
            rename_visia_files(Imp.validated_path, Imp.studynumber) 
        else:
            print("Regex match not matching for visia or std mask; check filename again")
            sys.exit(1)

    # .............Missing file handler.......................................

    filenamelist = Util.getAllFiles(Imp.validated_path, "*.jpg", depth=-1)
    dummy_list = CFS.replace_missing_barcodes_with_dummy(filenamelist, is_visia)

    # if missing files found
    if dummy_list is not None:
        filenamelist = Util.getAllFiles(Imp.validated_path, "*.jpg", depth=-1) # read again to include dummy images

        # logging cloned files to handle missing barcodes
        
        with open(os.path.join(Imp.validated_path, Imp.dummy_log_file), "w") as log_f:
            for item in dummy_list:
                log_f.write(item + "\n")
                
    #final data cleaning - remove file counter and delete duplicate jpg images
    CFS.remove_redundant_jpgs(Imp.validated_path)
    CFS.remove_counter_from_filenames(Imp.validated_path)
    CFS.remove_bitmaps_from_tif_conversion(Imp.validated_path)
    
    
    #................... Excel layout template creation..........................
    
    dict_code = CFS.create_code_dict_for_excel_table(Imp.validated_path)
    
    area_code_list, time_code_list, \
    area_sorted, time_sorted, sub_sorted, \
    filepaths, filenamelist, \
    Main_mapping_dict, \
    transfer_list = create_excel(Imp.validated_path, 
                                    Imp.excelfile, 
                                    dict_code, 
                                    Imp.filename_mask)


    # console output - excel process confirmation
    open_path = os.getcwd()
    proc_excel = subprocess.Popen(f"start {os.path.join(open_path, Imp.excelfile)}", shell=True)
    proc_excel.wait()
    input("folder open..Press enter")

    # after user says 'y'
    wb = load_workbook(Imp.excelfile)
    # opens modified worksheet
    ws_modify_template = wb[MODIFY_TEMPLATE_SHEET]
    # reading cell A1 - only for CustomTemplate.py
    cell_a1 = ws_modify_template.cell(row=1, column=1).value  # checking if a * or § is found in this cell
    print(cell_a1)

    # condition handling column names for each row
    column_row_name_list = []
    for coln, col in enumerate(ws_modify_template.columns):
        for cell in col:
            if cell.value == COLUMN_NAMES_FLAG:
                print("found")
                special_col_index = utils.get_column_letter(cell.column)

                column_row_name_list = []
                for row in ws_modify_template[special_col_index]:
                    if row.value == "None" or row.value is None or row.value == "":
                        break
                    column_row_name_list.append(row.value)
                column_row_name_list = column_row_name_list[1:]
                print(column_row_name_list)
                break
        if coln == COLUMN_SCAN_LIMIT:
            break

    if cell_a1 == CUSTOM_TEMPLATE_FLAG:

        rownames, colnames = CFS.get_row_and_columns(ws_modify_template, 3, 2)

        randomized_files_iterator = None
        if Imp.randomfilepath is not None and Imp.isVisia:
            try:
                randomized_files_iterator = randomize_custom_rand_alllightingcode(filepaths,
                                                          Imp.randomfilepath,
                                                          area_sorted, time_sorted, sub_sorted)
            except IndexError:
                print("Custom - file List index out of range; Check the random file again for duplicates or Missing values")
                sys.exit(1)
        elif Imp.randomfilepath is not None and not Imp.isVisia:
            try:
                randomized_files_iterator = randomize_std(filepaths,
                                                          Imp.randomfilepath,
                                                          area_sorted, time_sorted, sub_sorted)
            except IndexError:
                print("Custom - file List index out of range; Check the random file again for duplicates or Missing values")
                sys.exit(1)
        

        cst = Custom(rownames, colnames)
        tex_file = cst.get_final_tex_file(sub_sorted, area_sorted, time_sorted,
                                          rownames, colnames, Main_mapping_dict, randomized_files_iterator, Imp)
        
    else:
        rownames, colnames = CFS.get_row_and_columns(ws_modify_template, 2, 1)
        if time_sorted[0] in rownames[0]:
            if Imp.randomfilepath is not None:
                try: 
                    randomized_files_iterator, filepathrandlist = randomize_transp(filepaths, 
                                                                                   Imp.randomfilepath, 
                                                                                   area_sorted, time_sorted, sub_sorted)
                    
                    path_sep_list = [path[0] for path in filepathrandlist] # retrievs paths from the tuple
                    name_sep_file_list = [str_path.split("\\")[-1] for str_path in path_sep_list] # retrieves names from paths

                    tps = Transpose(area_code_list, time_code_list, 
                                                area_sorted, time_sorted,
                                                rownames,
                                                transfer_list, Imp)
                    
                    tex_file = tps.get_final_tex_file(sub_sorted, area_sorted, time_sorted,
                                                                                        rownames, colnames,
                                                                                        filepaths, filenamelist, 
                                                                                        randomized_files_iterator, Imp)
                except IndexError:
                    print("List index out of range; Check the random file again for duplicates or Missing values")
                    sys.exit(1)
            else:
                randomized_files_iterator = None

                tps = Transpose(area_code_list, time_code_list, 
                                            area_sorted, time_sorted,
                                            rownames,
                                            transfer_list, Imp)
                
                tex_file = tps.get_final_tex_file(sub_sorted, area_sorted, time_sorted,
                                    rownames, colnames,
                                    filepaths, filenamelist, 
                                    randomized_files_iterator, Imp)

        else:
            if Imp.randomfilepath is not None:
                try: 
                    randomized_files_iterator = randomize_std(filepaths, 
                                                              Imp.randomfilepath, 
                                                              area_sorted, time_sorted, sub_sorted)
                except IndexError:
                    print("List index out of range; Check the random file again for duplicates or Missing values")
                    sys.exit(1)
            else:
                randomized_files_iterator = None

            std = Standard(area_code_list, time_code_list, rownames)
            tex_file = std.get_final_tex_file(sub_sorted, area_sorted, time_sorted,
                                              rownames, colnames, 
                                        filepaths, filenamelist, 
                                        randomized_files_iterator, 
                                        Imp)
            
    # ........................ Additional features/study specific layouts .................
    if len(column_row_name_list) > 0:
        tex_file = give_each_row_columnames(column_row_name_list, tex_file)


    final_tex_file = Insert_description_file(tex_file)

    # ................. Generate the PDF from tex file automatically and display it to the user .................
    if os.path.isfile(final_tex_file):
        try:
            proc1 = subprocess.Popen(f"pdflatex -interaction=nonstopmode -halt-on-error {final_tex_file}  && pause", shell=True)
            proc1.wait()
            print("All files generated...starting archiving and cleaning process")
            sleep(1)
        except (OSError, subprocess.SubprocessError):
            os.startfile(final_tex_file)
            input("Press enter to exit...")
    
    CFS.archive_data(Imp.validated_path, 
                     Imp.studynumber, 
                     Imp.Test_type, 
                     Imp.file_extension)

# ....................................................................................................
# ............................THE END........THE END..........THE END........THE END.......................................
# .........................................................................................................................
if __name__ == '__main__':
    main()
