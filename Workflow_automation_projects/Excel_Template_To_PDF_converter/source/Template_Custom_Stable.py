from openpyxl.reader.excel import load_workbook
from openpyxl import load_workbook

import Internal_Imports_Stable as Imp 
from Common_Functions_Stable import create_Latex_document
from Latex_File_Create_Stable import create_final_latex_file


def customize(area_sorted, time_sorted, sub_sorted, rownames, colnames, Main_mapping_dict, random_iter, Imp):
    """
    Users can customize layout as per their own preferences and study needs
    """
    next_label_count = "".join(rownames).count("next")

    
    # Reads the modified template
    wb = load_workbook(Imp.excelfile)
    ws4 = wb['Modify_Template_here']

    nr_col=len(colnames)#len(time_list)
    nr_row=len(rownames)#len(area_list)

    max_height=round(1.0/(1.15*nr_row), 2)
    if next_label_count > 0:
        max_height=round((1.0*next_label_count)/(1.25*nr_row), 2)
    
    max_width=round(1.0/(1.2*nr_col), 2)
        
    lax_document, new_page_alert = create_Latex_document(colnames, rownames, sub_sorted,  max_height, max_width, 
                                                         area_sorted, time_sorted, ws4, Main_mapping_dict, Type="Custom", RandomList=random_iter)


    custom_tex_file = create_final_latex_file(Imp.studynumber, Imp.header, Imp.pagestyle, Imp.hypersetup, 
                                              lax_document, new_page_alert, colnames, Imp.Test_type, draft=Imp.draft_flag)
    
    return custom_tex_file