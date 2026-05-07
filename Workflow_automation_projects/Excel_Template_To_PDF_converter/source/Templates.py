
from Latex_File_Create_Stable import create_final_latex_file
from Common_Functions_Stable import create_Latex_document

from openpyxl.reader.excel import load_workbook
from openpyxl import load_workbook



class Standard:
    
    def __init__(self, area_code_list, time_code_list, rownames):

        self.max_height, self.max_width = self.get_max_height_and_width(area_code_list, time_code_list, rownames)

    def get_max_height_and_width(self, area_code_list, time_code_list, rownames):
        
        nr_row = len(area_code_list)
        nr_col = len(time_code_list)
        
        max_height = round(1.0/(1.25*nr_row), 2) 
        next_label_count = "".join(rownames).count("next")
        if next_label_count > 0:
            max_height= round((1.0*next_label_count)/(1.25*nr_row), 2)
        max_width = round(1.0/(1.5*nr_col), 2) 
        return max_height, max_width
    
    def get_final_tex_file(self, 
                           sub_sorted, area_sorted, time_sorted,
                           rownames, colnames, 
                           filepaths, filenamelist, 
                           random_iter, Imp,
                           Type="Standard"):
        
        lax_document, new_page_alert = create_Latex_document(colnames,  rownames,
                                                    sub_sorted, 
                                                    self.max_height, self.max_width, 
                                                    area_sorted, time_sorted, 
                                                    filepaths, filenamelist, 
                                                    Type, 
                                                    RandomList=random_iter)
        
        tex_file = create_final_latex_file(Imp.studynumber, Imp.header, Imp.pagestyle, Imp.hypersetup, 
                                                lax_document, new_page_alert, 
                                                colnames, Imp.Test_type, 
                                                draft=Imp.draft_flag)
        
        return tex_file
    

class Transpose(Standard):
    def __init__(self, area_code_list, time_code_list, 
                                area_sorted, time_sorted, 
                                rownames,
                                transfer_list, Imp, random=True): # TODO: NOTE random=True condition had previously the same function call for if and else true/false
        
        #swap area_code_list and time_code_list for nrows and ncolumns
        area_code_list_swapped = time_code_list
        time_code_list_swapped = area_code_list
        
        self.max_height, self.max_width = super().get_max_height_and_width(area_code_list_swapped, time_code_list_swapped,
                                                                            rownames)

        self.transpose_excel_template(area_code_list, time_code_list, 
                                            area_sorted, time_sorted,
                                            transfer_list, Imp)

        
    def transpose_excel_template(self, area_code_list, time_code_list, 
                                        area_sorted, time_sorted, 
                                        transfer_list, Imp):
        
        wb = load_workbook(Imp.excelfile)
        ws4 = wb['Modify_Template_here']
        
        transfer_list_transposed = []
        for tcode in time_sorted:
            for acode in area_sorted:
                for tvalue in transfer_list:
                    if acode in tvalue and tcode in tvalue:
                        transfer_list_transposed.append(tvalue)

        Transpose_Iter = iter(transfer_list_transposed) # initialize iterator

        for ttn, _ in enumerate(time_code_list, start=2):   # Replacing the item in rows/columns  in modified excel sheet with the derandomised value for the first subject
            for aan, _ in enumerate(area_code_list, start=2): # TODO: For custom template - value in the rows and col must be read first and then rand value must be swapped with it
                ws4.cell(row=ttn, column=aan).value = next(Transpose_Iter)

        wb.save(Imp.excelfile)
        
    def get_final_tex_file(self, 
                           sub_sorted, area_sorted, time_sorted,
                           rownames, colnames, 
                           filepaths, filenamelist, 
                           random_iter, Imp,
                           Type="Standard"):
        
        # Notice that here ara_sorted and time_sorted are swapped while passing it to this function
        # this is different to the standard passing of arguments
        lax_document, new_page_alert = create_Latex_document(colnames,  rownames,
                                                    sub_sorted, 
                                                    self.max_height, self.max_width, 
                                                    time_sorted, area_sorted,
                                                    filepaths, filenamelist, 
                                                    Type, 
                                                    RandomList=random_iter)
        
        tex_file = create_final_latex_file(Imp.studynumber, Imp.header, Imp.pagestyle, Imp.hypersetup, 
                                                lax_document, new_page_alert, 
                                                colnames, Imp.Test_type, 
                                                draft=Imp.draft_flag)
        
        return tex_file                     

class Custom(Standard):
    def __init__(self, rownames, colnames): # TODO: NOTE random=True condition had previously the same function call for if and else true/false
        
        #TODO: NOTE For custom template: row_names, colnames replaces   area_code_list, time_code_listarguments while calling this function
        self.max_height, self.max_width = self.get_max_height_and_width(rownames, colnames)
    
    
    def get_max_height_and_width(self, rownames, colnames):
        next_label_count = "".join(rownames).count("next")
        nr_col=len(colnames)#len(time_list)
        nr_row=len(rownames)#len(area_list)

        max_height=round(1.0/(1.15*nr_row), 2)
        if next_label_count > 0:
            max_height=round((1.0*next_label_count)/(1.25*nr_row), 2)
        
        max_width=round(1.0/(1.2*nr_col), 2)
        return max_height, max_width
    
    #TODO: NOTE  ws4, main_mapping_dict replaces  filepaths, filenamelist arguments while calling this function        
    def get_final_tex_file(self, sub_sorted, area_sorted, time_sorted,
                                        rownames, colnames,
                                        main_mapping_dict, 
                                        random_iter, Imp):
        
        ws4 = load_workbook(Imp.excelfile)['Modify_Template_here']
        # Notice that here ara_sorted and time_sorted are swapped while passing it to this function
        # this is different to the standard passing of arguments
        lax_document, new_page_alert = create_Latex_document(colnames,  rownames,
                                                    sub_sorted, 
                                                    self.max_height, self.max_width, 
                                                    area_sorted, time_sorted,
                                                    ws4, main_mapping_dict, 
                                                    Type="Custom", 
                                                    RandomList=random_iter)
        
        tex_file = create_final_latex_file(Imp.studynumber, Imp.header, Imp.pagestyle, Imp.hypersetup, 
                                                lax_document, new_page_alert, 
                                                colnames, Imp.Test_type, 
                                                draft=Imp.draft_flag)
        
        return tex_file