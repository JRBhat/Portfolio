"""
Protocol Scanner  v2
====================
Reads an Excel file with study records, scans protocol folders for PDF/Word documents
containing "protocol" in the filename, extracts paragraphs referencing "foto"/"image"
and "image analysis", and writes results to a new Excel file.

Changes from v1:
  - Fixed Windows-unsafe timestamp in default output filename (colons are illegal)
  - Sentences are split once per document instead of twice (was called per extract_context)
  - Rows are processed in parallel with ThreadPoolExecutor (I/O bound speedup)
  - PDF library availability is probed once at startup, not on every call
  - Removed unused `pathlib.Path` import

Dependencies:
    pip install pandas openpyxl pymupdf python-docx
    (pymupdf is the fastest PDF text extractor; fallback to pdfplumber if unavailable)

Usage:
    python protocol_scanner_v2.py <input_excel.xlsx> [output_excel.xlsx]
"""

import os
import re
import sys
import logging
import warnings
from concurrent.futures import ThreadPoolExecutor, as_completed

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

log = logging.getLogger(__name__)

# Suppress noisy warnings from PDF/docx libs only
warnings.filterwarnings("ignore", category=DeprecationWarning, module="docx")
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

# ── Config (edit these if running directly rather than via command line) ───────
INPUT_EXCEL  = os.environ.get("INPUT_EXCEL", "")
OUTPUT_EXCEL = os.environ.get("OUTPUT_EXCEL", "")

# ── Constants ──────────────────────────────────────────────────────────────────
FOTO_PATTERN = re.compile(r"\b(photo|image)\b", re.IGNORECASE)
IMAGE_ANALYSIS_PATTERN = re.compile(r"\bimage\b.{0,60}\banalysis\b", re.IGNORECASE)
CONTEXT_SENTENCES = 5

# ── Probe PDF library availability once at startup ─────────────────────────────
try:
    import fitz as _fitz  # PyMuPDF
    _PDF_BACKEND = "fitz"
except ImportError:
    _fitz = None
    try:
        import pdfplumber as _pdfplumber
        _PDF_BACKEND = "pdfplumber"
    except ImportError:
        _pdfplumber = None
        _PDF_BACKEND = None
        log.error("Neither PyMuPDF nor pdfplumber is installed. PDF extraction will fail.")

log.debug("PDF backend: %s", _PDF_BACKEND)


# ── Text extraction helpers ────────────────────────────────────────────────────

def extract_text_pdf(path: str) -> str:
    """Extract all text from a PDF using the available backend."""
    if _PDF_BACKEND == "fitz":
        try:
            text_parts = []
            with _fitz.open(path) as doc:
                for page in doc:
                    text_parts.append(page.get_text())
            text = "\n".join(text_parts)
            log.debug("  [PDF/fitz] extracted %d chars from '%s'", len(text), path)
            return text
        except Exception as e:
            log.warning("  [PDF/fitz] error on '%s': %s", path, e)

    elif _PDF_BACKEND == "pdfplumber":
        try:
            text_parts = []
            with _pdfplumber.open(path) as pdf:
                for page in pdf.pages:
                    t = page.extract_text()
                    if t:
                        text_parts.append(t)
            text = "\n".join(text_parts)
            log.debug("  [PDF/pdfplumber] extracted %d chars from '%s'", len(text), path)
            return text
        except Exception as e:
            log.warning("  [PDF/pdfplumber] error on '%s': %s", path, e)

    else:
        log.error("  No PDF backend available.")
    return ""


def extract_text_docx(path: str) -> str:
    """Extract all text from a .doc/.docx file using python-docx."""
    try:
        from docx import Document
        doc = Document(path)
        text = "\n".join(p.text for p in doc.paragraphs)
        log.debug("  [DOCX] extracted %d chars from '%s'", len(text), path)
        return text
    except Exception as e:
        log.warning("  [DOCX] error on '%s': %s", path, e)
    return ""


# ── Context extraction ─────────────────────────────────────────────────────────

def split_sentences(text: str) -> list[str]:
    """Naive but fast sentence splitter."""
    sentences = re.split(r"(?<=[.!?])\s+", text.replace("\n", " "))
    return [s.strip() for s in sentences if s.strip()]


def extract_intro(text: str, n_lines: int = 50) -> str:
    """Return the first n_lines non-empty lines of text (intro / summary section)."""
    lines = [line for line in text.splitlines() if line.strip()]
    return "\n".join(lines[:n_lines])


def _context_from_sentences(sentences: list[str], pattern: re.Pattern, n: int = CONTEXT_SENTENCES) -> str:
    """
    Core context extractor — operates on a pre-split sentence list.
    Find every sentence matching *pattern*, collect ±n surrounding sentences,
    merge overlapping windows, and return passages joined by a separator.
    """
    total = len(sentences)
    if total == 0:
        return ""

    hit_indices = [i for i, s in enumerate(sentences) if pattern.search(s)]
    log.debug("  Sentences matching '%s': %d", pattern.pattern, len(hit_indices))

    if not hit_indices:
        return ""

    # Merge overlapping windows
    windows: list[tuple[int, int]] = []
    for idx in hit_indices:
        start = max(0, idx - n)
        end = min(total - 1, idx + n)
        if windows and start <= windows[-1][1] + 1:
            windows[-1] = (windows[-1][0], max(windows[-1][1], end))
        else:
            windows.append((start, end))

    passages = [" ".join(sentences[s:e + 1]) for s, e in windows]
    result = "\n\n--- [match] ---\n\n".join(passages)
    log.debug("  Extracted %d passage(s), total chars: %d", len(passages), len(result))
    return result


def extract_context(text: str, pattern: re.Pattern = FOTO_PATTERN, n: int = CONTEXT_SENTENCES) -> str:
    """Split text into sentences then delegate to _context_from_sentences."""
    return _context_from_sentences(split_sentences(text), pattern, n)


# ── Protocol file discovery ────────────────────────────────────────────────────

def find_protocol_file(folder: str) -> str | None:
    """
    Walk folder for a file whose name contains 'protocol' and is PDF or Word.
    PDFs are preferred; Word documents are fallback.
    Shortest path wins when multiple candidates exist.
    """
    if not os.path.isdir(folder):
        log.warning("  Folder does not exist or is not accessible: '%s'", folder)
        return None

    pdf_hits = []
    doc_hits = []

    for root, _dirs, files in os.walk(folder):
        for fname in files:
            lower = fname.lower()
            if "protocol" not in lower:
                continue
            ext = os.path.splitext(lower)[1]
            full = os.path.join(root, fname)
            if ext == ".pdf":
                pdf_hits.append(full)
            elif ext in (".doc", ".docx"):
                doc_hits.append(full)

    log.debug("  PDF hits: %s", pdf_hits)
    log.debug("  Word hits: %s", doc_hits)

    for hits in (pdf_hits, doc_hits):
        if hits:
            hits.sort(key=lambda p: (len(p), p))
            return hits[0]
    return None


# ── Per-row processing ─────────────────────────────────────────────────────────

def process_row(study_nr: str, foto_path: str, protocol_folder: str) -> dict:
    """Process a single study row and return result dict."""
    log.info("Processing StudyNr: %s", study_nr)
    log.debug("  FotoQualiSheet: %s", foto_path)
    log.debug("  ProtocolFolder: %s", protocol_folder)

    result = {
        "StudyNr": study_nr,
        "FotoQualiSheet path": foto_path,
        "ProtocolPath": protocol_folder,
        "Protocol": "",
        "Intro": "",
        "ImageAnalysis": "",
        "FotoTextDump": "",
    }

    proto_file = find_protocol_file(str(protocol_folder))
    if proto_file is None:
        log.warning("  No protocol file found in '%s'", protocol_folder)
        result["FotoTextDump"] = "[No protocol file found]"
        return result

    log.info("  Found protocol file: %s", proto_file)
    result["Protocol"] = proto_file

    ext = os.path.splitext(proto_file)[1].lower()
    if ext == ".pdf":
        text = extract_text_pdf(proto_file)
    elif ext in (".doc", ".docx"):
        text = extract_text_docx(proto_file)
    else:
        log.warning("  Unsupported file type: %s", ext)
        result["FotoTextDump"] = "[Unsupported file type]"
        return result

    if not text:
        log.warning("  No text extracted from '%s'", proto_file)
        result["FotoTextDump"] = "[No text extracted from protocol file]"
        return result

    result["Intro"] = extract_intro(text)

    # Split sentences once and reuse for both pattern searches
    sentences = split_sentences(text)
    log.debug("  Total sentences: %d", len(sentences))

    image_analysis = _context_from_sentences(sentences, IMAGE_ANALYSIS_PATTERN)
    if image_analysis:
        log.info("  Image analysis context found (%d chars)", len(image_analysis))
        result["ImageAnalysis"] = image_analysis
    else:
        log.info("  No 'image analysis' references found in '%s'", proto_file)
        result["ImageAnalysis"] = "[No image analysis references found]"

    foto_context = _context_from_sentences(sentences, FOTO_PATTERN)
    if foto_context:
        log.info("  Foto/image context found (%d chars)", len(foto_context))
        result["FotoTextDump"] = foto_context
    else:
        log.info("  No foto/image references found in '%s'", proto_file)
        result["FotoTextDump"] = "[No foto/image references found]"

    return result


# ── Excel output ───────────────────────────────────────────────────────────────

def write_output_excel(records: list[dict], output_path: str) -> None:
    """Write results to a formatted Excel file."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Results"

    headers = ["StudyNr", "FotoQualiSheet path", "ProtocolPath", "Protocol", "Intro", "ImageAnalysis", "FotoTextDump"]

    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="2E4057")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    ws.row_dimensions[1].height = 30

    data_font = Font(name="Arial", size=10)
    wrap_align = Alignment(vertical="top", wrap_text=True)

    for row_idx, rec in enumerate(records, 2):
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=rec.get(h, ""))
            cell.font = data_font
            cell.alignment = wrap_align

    col_widths = [20, 45, 45, 45, 80, 80, 80]
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    wb.save(output_path)
    log.info("Output written to '%s'", output_path)


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    timestamp = pd.Timestamp.now().strftime("%Y-%m-%d_%H-%M-%S")

    if len(sys.argv) >= 2:
        input_path  = sys.argv[1]
        output_path = sys.argv[2] if len(sys.argv) > 2 else f"protocol_insights_{timestamp}.xlsx"
    elif INPUT_EXCEL.strip():
        input_path  = INPUT_EXCEL
        output_path = OUTPUT_EXCEL.strip() or f"protocol_insights_{timestamp}.xlsx"
    else:
        print("Usage: python protocol_scanner_v2.py <input.xlsx> [output.xlsx]")
        print("       — or set INPUT_EXCEL / OUTPUT_EXCEL at the top of the script.")
        sys.exit(1)

    log.info("Reading input file: '%s'", input_path)
    try:
        df = pd.read_excel(input_path, dtype=str)
    except Exception as e:
        log.error("Failed to read input Excel: %s", e)
        sys.exit(1)

    df.columns = [c.strip() for c in df.columns]
    log.info("Columns found: %s", list(df.columns))
    log.info("Rows to process: %d", len(df))

    required_cols = ["StudyNr", "FotoQualiSheet path", "ProtocolPath"]
    for col in required_cols:
        if col not in df.columns:
            log.error("Missing expected column: '%s'. Found: %s", col, list(df.columns))
            sys.exit(1)

    tasks = []
    for _, row in df.iterrows():
        study_nr = str(row["StudyNr"]).strip()
        if not study_nr or study_nr.lower() == "nan":
            log.debug("  Skipping empty row")
            continue
        tasks.append((
            study_nr,
            str(row["FotoQualiSheet path"]).strip(),
            str(row["ProtocolPath"]).strip(),
        ))

    log.info("Submitting %d rows to thread pool …", len(tasks))

    records: list[dict | None] = [None] * len(tasks)
    with ThreadPoolExecutor(max_workers=min(8, len(tasks) or 1)) as pool:
        future_to_idx = {
            pool.submit(process_row, *args): idx
            for idx, args in enumerate(tasks)
        }
        for future in as_completed(future_to_idx):
            idx = future_to_idx[future]
            try:
                records[idx] = future.result()
            except Exception as exc:
                log.error("  Row %d raised an exception: %s", idx, exc)
                study_nr = tasks[idx][0]
                records[idx] = {"StudyNr": study_nr, "FotoTextDump": f"[Error: {exc}]"}

    log.info("Processing complete. Writing %d records to '%s'", len(records), output_path)
    write_output_excel(records, output_path)
    print(f"\nDone! Results saved to: {output_path}")
    print("Log saved to: protocol_scanner.log")


if __name__ == "__main__":
    logging.basicConfig(
        level=logging.DEBUG,
        format="%(asctime)s [%(levelname)s] %(message)s",
        handlers=[
            logging.StreamHandler(sys.stdout),
            logging.FileHandler("protocol_scanner.log", encoding="utf-8"),
        ],
    )
    main()
