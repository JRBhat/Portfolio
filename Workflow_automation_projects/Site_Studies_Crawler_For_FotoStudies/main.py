"""
Search clinical study directories for Foto qualification/Laufzettel files.
Uses PowerShell (via subprocess) for fast file system traversal on Windows.
Outputs results to Excel.
"""

import subprocess
import sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


def find_foto_files_powershell(root_dir: str) -> list[dict]:
    """
    Use PowerShell to rapidly scan for Foto files in .docu subfolders.
    PowerShell's Get-ChildItem is significantly faster than pure Python os.walk
    for large directory trees on Windows (leverages native filesystem APIs).
    """
    # Force PowerShell to output UTF-8, then decode as UTF-8 on Python side
    ps_command = f"""
[Console]::OutputEncoding = [System.Text.Encoding]::UTF8
$RootDir = '{root_dir}'
Get-ChildItem -Path $RootDir -Directory -Recurse | Where-Object {{
    $_.Name -imatch '^\d{{2}}\.\d{{4}}-\d{{2}}_'
}} | ForEach-Object {{
    $studyFolder = $_
    $studyNum = $studyFolder.Name
    $prefix = ($studyNum -split '-')[0]
    $protocolFolder = Join-Path $studyFolder.FullName "$prefix.protocol"

    $docuFolder = Get-ChildItem -Path $studyFolder.FullName -Directory -Recurse |
        Where-Object {{ $_.Name -imatch '\.docu$' }} |
        Select-Object -First 1

    if ($docuFolder) {{
        $matchedFile = Get-ChildItem -Path $docuFolder.FullName -File |
            Where-Object {{ $_.Name -imatch 'foto' -and ($_.Name -imatch 'quali' -or $_.Name -imatch 'laufzettel') }} |
            Select-Object -First 1

        if ($matchedFile) {{
            "$studyNum|$($matchedFile.FullName)|$protocolFolder"
        }}
    }}
}}
"""
    result = subprocess.run(
        ["powershell", "-NoProfile", "-NonInteractive", "-Command", ps_command],
        capture_output=True,   # get raw bytes
    )
    # Decode stdout as UTF-8; fall back to cp1252 (German Windows default) on error
    try:
        stdout = result.stdout.decode("utf-8")
    except UnicodeDecodeError:
        stdout = result.stdout.decode("cp1252", errors="replace")

    if result.returncode != 0 and result.stderr:
        stderr = result.stderr.decode("utf-8", errors="replace")
        print(f"[PowerShell warning/error]: {stderr.strip()}", file=sys.stderr)

    rows = []
    for line in stdout.strip().splitlines():
        line = line.strip()
        if not line or "|" not in line:
            continue
        parts = line.split("|", 2)
        if len(parts) == 3:
            rows.append({
                "StudyNr": parts[0],
                "FotoQualiSheet path": parts[1],
                "ProtocolPath": parts[2],
            })
    return rows


def write_excel(rows: list[dict], output_path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Results"

    headers = ["StudyNr", "FotoQualiSheet path", "ProtocolPath"]
    header_fill = PatternFill(fill_type="solid", fgColor="4472C4")
    header_font = Font(name="Arial", bold=True, color="FFFFFF")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    data_font = Font(name="Arial", size=10)
    for row_idx, row in enumerate(rows, 2):
        ws.cell(row=row_idx, column=1, value=row["StudyNr"]).font = data_font
        ws.cell(row=row_idx, column=2, value=row["FotoQualiSheet path"]).font = data_font
        ws.cell(row=row_idx, column=3, value=row["ProtocolPath"]).font = data_font

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 80
    ws.column_dimensions["C"].width = 70

    wb.save(output_path)
    print(f"Results saved to: {output_path}")


def main():
    if len(sys.argv) < 2:
        print("Usage: python search_clinical_studies.py <root_directory> [output.xlsx]")
        print("Example: python search_clinical_studies.py 'C:\\Studies' results.xlsx")
        sys.exit(1)

    root_dir = sys.argv[1]
    output_xlsx = sys.argv[2] if len(sys.argv) > 2 else "clinical_study_results.xlsx"

    if not Path(root_dir).exists():
        print(f"Error: Directory '{root_dir}' does not exist.")
        sys.exit(1)

    print(f"Scanning: {root_dir}")
    rows = find_foto_files_powershell(root_dir)

    if not rows:
        print("No matching files found.")
    else:
        print(f"Found {len(rows)} matching file(s).")

    write_excel(rows, output_xlsx)


if __name__ == "__main__":
    main()