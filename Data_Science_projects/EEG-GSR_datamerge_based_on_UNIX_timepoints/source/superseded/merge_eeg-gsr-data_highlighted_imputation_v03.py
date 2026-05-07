import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def merge_device_datasets(d1path, d2path, output_path):
    # Load datasets
    d1 = pd.read_excel(d1path)  # Expected columns: timepoint, value1, etc.
    d2 = pd.read_csv(d2path, header=None, skiprows=3, sep="\t", usecols=[0,1,2,3,4])  # Expected columns: timepoint, value2, etc.
    
    # Set proper column headers
    d1.columns = ['ch1', 'ch2', 'ch3', 'ch4', 'ch5', 'ch6', 'ch7', 'ch8', 'X', 'Y', 'Z', 'marker', 'timestamp']
    new_headers = ['timestamp', 'GSR_range', 'Skin_conductance_uS', 'Skin_resistence_kOhms', 'PPG_mV']
    d2.columns = new_headers
    
    # Convert Unix-based timepoints (in milliseconds) to datetime and sort
    d1["timestamp"] = pd.to_datetime(d1["timestamp"], unit="ms")
    d2["timestamp"] = pd.to_datetime(d2["timestamp"], unit="ms")
    d1 = d1.sort_values("timestamp")
    d2 = d2.sort_values("timestamp")
    
    # Merge the datasets using an outer join to preserve all timestamps
    merged_outer = pd.merge(d1, d2, on="timestamp", how="outer", sort=True)
    merged_outer.sort_values("timestamp", inplace=True)
    
    # Save a mask of where missing values exist BEFORE interpolation.
    missing_mask = merged_outer.isna()
    
    # Define columns by type
    eeg_cols = ['ch1', 'ch2', 'ch3', 'ch4', 'ch5', 'ch6', 'ch7', 'ch8', 'X', 'Y', 'Z']
    gsr_cols = ['GSR_range', 'Skin_conductance_uS', 'Skin_resistence_kOhms']
    ppg_col = 'PPG_mV'
    
    # Fill missing values:
    # - For EEG channels and PPG data, use linear interpolation.
    merged_outer[eeg_cols] = merged_outer[eeg_cols].interpolate(method="linear")
    merged_outer[ppg_col] = merged_outer[ppg_col].interpolate(method="linear")
    
    # - For GSR values, use forward-fill (ffill) since these change gradually.
    merged_outer[gsr_cols] = merged_outer[gsr_cols].fillna(method="ffill")
    
    # - For any remaining NaNs (e.g., at the beginning), back-fill.
    merged_outer.fillna(method="bfill", inplace=True)
    
    # Save the merged dataset with interpolated values to Excel
    excel_output = output_path  # e.g., r"H:\merged_outer.xlsx"
    merged_outer.to_excel(excel_output, index=False)
    
    # Load the workbook with openpyxl to apply formatting.
    wb = load_workbook(excel_output)
    ws = wb.active
    
    # Define a yellow fill for highlighting
    highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    # Get DataFrame column names (these match the Excel header in row 1)
    cols = merged_outer.columns.tolist()
    
    # Iterate over the DataFrame cells (Excel rows start at 2, since row 1 is header)
    for r_idx, _ in enumerate(merged_outer.itertuples(index=False), start=2):
        for c_idx, col in enumerate(cols, start=1):
            # If the original missing mask was True at this cell, apply the highlight.
            if missing_mask.loc[merged_outer.index[r_idx-2], col]:
                ws.cell(row=r_idx, column=c_idx).fill = highlight_fill
                
    # Save the workbook with highlighted changes
    wb.save(excel_output)
    print(f"Merged and interpolated dataset saved with highlighted filled cells as '{excel_output}'.")

def main():
    eeg = "data/eeg_export/sample_eeg_main.xlsx"
    gsr = "data/eeg_export/eeg_study_Session23_Shimmer_A5C7_Calibrated_PC.csv"
    output_excel = r"H:\merged_outer_v3.xlsx"
    merge_device_datasets(eeg, gsr, output_excel)
    
if __name__ == "__main__":
    main()
