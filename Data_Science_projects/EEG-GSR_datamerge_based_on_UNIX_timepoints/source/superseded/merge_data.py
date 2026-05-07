import pandas as pd
import numpy as np
import scipy.signal as signal
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import os

# from calculate_bpm import calculate_bpm_from_ppg
from calculate_bpm_best_practices import calculate_bpm_from_ppg



def merge_device_datasets(d1path, d2path, Path):
    """Merge EEG and GSR datasets, replacing PPG_mV with BPM values, fill missing values, and highlight the imputed cells."""
    
    # Save final dataset to Excel with highlighting
    subj_name = Path.split("\\")[-1]
    output_excel = os.path.join(Path, rf"merged_highlighted_{subj_name}_bpmcorrected.xlsx")  # --- Addition: Excel output path
    output_csv = os.path.join(Path, rf"merged_clean_{subj_name}_bpmcorrected.csv") 
    if not os.path.exists(output_excel) and not os.path.exists(output_csv):
        
        # Load EEG dataset (expected to contain timepoint and EEG values)
        d1 = pd.read_csv(d1path, sep="\t") 
        
        # Load GSR dataset, adjusting headers
        d2 = pd.read_csv(d2path, header=None, skiprows=3, sep="\t", usecols=[0,1,2,3,4])  

        # Set proper column headers
        d1.columns = ['ch1', 'ch2', 'ch3', 'ch4', 'ch5', 'ch6', 'ch7', 'ch8', 'X', 'Y', 'Z', 'marker', 'timestamp']
        new_headers = ['timestamp', 'GSR_range', 'Skin_conductance_uS', 'Skin_resistence_kOhms', 'PPG_mV']
        d2.columns = new_headers
        
        # Convert timestamps from Unix milliseconds to datetime for merging purposes
        d1["timestamp"] = pd.to_datetime(d1["timestamp"], unit="ms") + pd.Timedelta(hours=2)
        d2["timestamp"] = pd.to_datetime(d2["timestamp"], unit="ms") + pd.Timedelta(hours=2)
        
        # Sort data by timestamp
        d1 = d1.sort_values("timestamp")
        d2 = d2.sort_values("timestamp")
        
        # Create a copy of d2 for BPM calculation with numeric timestamps (in ms)
        d2_for_bpm = d2.copy()
        # Convert d2_for_bpm timestamps back to numeric (ms) for BPM calculation
        d2_for_bpm["timestamp"] = d2_for_bpm["timestamp"].astype(np.int64) // 10**6  # --- Correction: integer division to get ms
        
        # Sort d2_for_bpm by timestamp (numeric)
        d2_for_bpm = d2_for_bpm.sort_values("timestamp")
        
        # Compute BPM from PPG signal using the copy with numeric timestamps
        bpm_df = calculate_bpm_from_ppg(d2_for_bpm)

        # Merge EEG and GSR datasets using an outer join
        merged_outer = pd.merge(d1, d2, on="timestamp", how="outer", sort=True)

        # Merge BPM results, replacing PPG_mV with BPM using a nearest match (tolerance 100ms)
        merged_final = pd.merge_asof(
            merged_outer, bpm_df, on="timestamp", direction="nearest", tolerance=pd.Timedelta("100ms")
        )

        # # Replace PPG_mV column with BPM values
        # merged_final["PPG_mV"] = merged_final["BPM"]
        # merged_final.drop(columns=["BPM"], inplace=True)  # Remove redundant BPM column

        # --- Addition Start: Interpolation & Highlighting ---
        # Save a mask of where missing values originally exist in the merged dataset
        missing_mask = merged_final.isna()
        
        # Define columns to fill:
        # EEG channels (columns starting with "ch" or in ["X", "Y", "Z"])
        eeg_cols = [col for col in merged_final.columns if col.startswith("ch") or col in ["X", "Y", "Z"]]
        # GSR columns
        gsr_cols = ["GSR_range", "Skin_conductance_uS", "Skin_resistence_kOhms"]
        # BPM column (replacing PPG_mV)
        bpm_col = "BPM"
        ppg_col = "PPG_mV"
        # Interpolate EEG channels, PPG and BPM using linear interpolation
        if eeg_cols:
            merged_final[eeg_cols] = merged_final[eeg_cols].interpolate(method="linear")
        merged_final[bpm_col] = merged_final[bpm_col].interpolate(method="linear")
        merged_final[ppg_col] = merged_final[ppg_col].interpolate(method="linear")
        
        # For GSR values, use forward-fill and then back-fill for any remaining gaps
        merged_final[gsr_cols] = merged_final[gsr_cols].fillna(method="ffill")
        merged_final.fillna(method="bfill", inplace=True)
        # --- Addition End ---
        
        # Save final dataset to Excel with highlighting
        merged_final.to_excel(output_excel, index=False)
        
        # --- Addition: Highlighting imputed cells in Excel ---
        wb = load_workbook(output_excel)
        ws = wb.active
        highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        
        cols = merged_final.columns.tolist()
        # Data starts from row 2 in Excel (row 1 is header)
        for r_idx, row in enumerate(merged_final.itertuples(index=False), start=2):
            for c_idx, col in enumerate(cols, start=1):
                # If the original cell was missing, highlight it
                if missing_mask.loc[merged_final.index[r_idx-2], col]:
                    ws.cell(row=r_idx, column=c_idx).fill = highlight_fill
                    
        wb.save(output_excel)
        print(f"Merged dataset with interpolated and highlighted cells saved as '{output_excel}'.")
        # Also save CSV if needed
        # merged_final.to_csv(os.path.join(Path, output_csv), index=False)
        # print(f"Merged dataset CSV saved as '{output_csv}'.")
        # --- End Highlighting Addition ---
        return output_excel
    else:
        print(f"{output_excel} and {output_csv} already exist.")
        print("Skipping merge and proceeding to FFA calculation...")
        return output_excel
    
def main():
    pass


    
if __name__ == "__main__":
    main()