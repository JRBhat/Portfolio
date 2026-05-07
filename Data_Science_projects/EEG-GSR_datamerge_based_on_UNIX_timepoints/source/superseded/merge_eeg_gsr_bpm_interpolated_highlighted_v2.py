import pandas as pd
import numpy as np
import scipy.signal as signal
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import os
from tests.diagnostics_heart_rate import diagnose_ppg


def calculate_bpm_from_ppg(df):
    """Process the PPG signal and compute BPM using peak detection and a rolling window approach."""
    
    diagnose_ppg(df)
    
    
     #————— FIXED TIME CONVERSION —————
    # 1) Force timestamp into datetime64 (if not already)
    df["timestamp"] = pd.to_datetime(df["timestamp"])                         # ← FIXED
    # 2) Save the base timestamp for later
    base_timestamp = df["timestamp"].iloc[0]                                  # ← FIXED
    # 3) Build a "time" vector in true seconds
    time = (df["timestamp"] - base_timestamp).dt.total_seconds()              # ← FIXED
    # ————————————————————————————————————
    
    # If you still need numeric timestamps for something else, you can get them via:
    # df["timestamp_ms"] = (df["timestamp"] - pd.Timestamp("1970-01-01")) \
    #                           .dt.total_seconds() * 1000
    
    # If the timestamp column is datetime, convert it to numeric (milliseconds)
    if pd.api.types.is_datetime64_any_dtype(df["timestamp"]):
        df["timestamp"] = df["timestamp"].astype(np.int64) / 1e6  
        print("Converted datetime timestamps to numeric (ms).")
    
    # Ensure timestamps are numeric (should now be in milliseconds)
    df["timestamp"] = pd.to_numeric(df["timestamp"], errors="coerce")
    
    # Save the base timestamp (first timestamp) as a datetime for later conversion
    base_timestamp = pd.to_datetime(df["timestamp"].iloc[0], unit="ms")  # --- Correction
    
    # Convert time differences from ms to seconds (relative time)
    time = (df["timestamp"] - df["timestamp"].iloc[0]) / 1000  # Now in seconds
    
    # Extract PPG signal
    ppg_signal = df["PPG_mV"]
    
    
    # Determine sampling frequency (Hz)
    dt = np.diff(time)
    fs = 1 / np.mean(dt)
    print("Sampling frequency (fs):", fs)
    
    # Desired cutoff frequencies in Hz
    desired_lowcut, desired_highcut = 0.5, 5
    nyquist = fs / 2.0
    if nyquist <= desired_lowcut:
        raise ValueError(f"Sampling frequency too low. Nyquist ({nyquist} Hz) must be greater than lowcut ({desired_lowcut} Hz).")
    
    # Normalize the cutoff frequencies
    low = desired_lowcut / nyquist
    high = desired_highcut / nyquist

    # Adjust if necessary to be safely within (0, 1)
    if low <= 0:
        low = 0.01  # minimal normalized frequency
        print("Adjusted normalized low cutoff to 0.01")
    if high >= 1:
        high = 0.99  # maximal normalized frequency
        print("Adjusted normalized high cutoff to 0.99")
    if high <= low:
        raise ValueError("Normalized high cutoff must be greater than normalized low cutoff.")
    
    print("Normalized low cutoff:", low, "Normalized high cutoff:", high)
    
    # Apply a bandpass filter with the adjusted cutoffs
    b, a = signal.butter(3, [low, high], btype='band')
    ppg_filtered = signal.filtfilt(b, a, ppg_signal)

    # Smooth the signal with a moving average filter
    window_size = int(fs * 0.5)
    if window_size < 1:
        window_size = 1  
    ppg_smoothed = np.convolve(ppg_filtered, np.ones(window_size) / window_size, mode='same')

    # Peak detection using an adaptive threshold (75th percentile)
    percentile_value = np.percentile(ppg_smoothed, 75)
    peaks, _ = signal.find_peaks(ppg_smoothed, height=percentile_value, distance=int(fs * 0.5))
    
    if len(peaks) == 0:
        print("No peaks detected in PPG signal.")
        return pd.DataFrame(columns=["timestamp", "BPM"])
    
    # Extract peak times (relative time in seconds)
    peak_times = time.iloc[peaks]

    # Compute BPM using a rolling 10-second window
    window_length = 10  
    bpm_values = []
    bpm_times = []

    for i in range(1, len(peak_times)):
        recent_peaks = peak_times[peak_times > (peak_times.iloc[i] - window_length)]
        if len(recent_peaks) > 1:
            ibi = np.diff(recent_peaks)  # Inter-beat intervals in seconds
            avg_ibi = np.mean(ibi)
            bpm = 60 / avg_ibi  # Convert to BPM
            bpm_values.append(bpm)
            bpm_times.append(peak_times.iloc[i])
    
    # Convert the relative peak times (in seconds) back to absolute datetime

    bpm_times = [base_timestamp + pd.to_timedelta(t, unit='s') for t in bpm_times]

    
    # Convert BPM results into a DataFrame
    bpm_df = pd.DataFrame({
        "timestamp": bpm_times,
        "BPM": bpm_values
    })
    
    return bpm_df


def merge_device_datasets(d1path, d2path, Path):
    """Merge EEG and GSR datasets, replacing PPG_mV with BPM values, fill missing values, and highlight the imputed cells."""
    
    # Load EEG dataset (expected to contain timepoint and EEG values)
    d1 = pd.read_csv(d1path, sep="\t") 
    
    # Load GSR dataset, adjusting headers
    d2 = pd.read_csv(d2path, header=None, skiprows=3, sep="\t", usecols=[0,1,2,3,4])  

    # Set proper column headers
    d1.columns = ['ch1', 'ch2', 'ch3', 'ch4', 'ch5', 'ch6', 'ch7', 'ch8', 'X', 'Y', 'Z', 'marker', 'timestamp']
    new_headers = ['timestamp', 'GSR_range', 'Skin_conductance_uS', 'Skin_resistence_kOhms', 'PPG_mV']
    d2.columns = new_headers
    
    # Convert timestamps from Unix milliseconds to datetime for merging purposes
    d1["timestamp"] = pd.to_datetime(d1["timestamp"], unit="ms")
    d2["timestamp"] = pd.to_datetime(d2["timestamp"], unit="ms")
    
    # Sort data by timestamp
    d1 = d1.sort_values("timestamp")
    d2 = d2.sort_values("timestamp")
    
    # Create a copy of d2 for BPM calculation with numeric timestamps (in ms)
    d2_for_bpm = d2.copy()
    # Convert d2_for_bpm timestamps back to numeric (ms) for BPM calculation
    d2_for_bpm["timestamp"] = d2_for_bpm["timestamp"].astype(np.int64) // 10**6  # --- Correction: integer division to get ms
    
    # Sort d2_for_bpm by timestamp (numeric)
    d2_for_bpm = d2_for_bpm.sort_values("timestamp")
    
    # Compute BPM from PPG signal using the copy with numeric timestamps
    bpm_df = calculate_bpm_from_ppg(d2_for_bpm)

    # Merge EEG and GSR datasets using an outer join
    merged_outer = pd.merge(d1, d2, on="timestamp", how="outer", sort=True)

    # Merge BPM results, replacing PPG_mV with BPM using a nearest match (tolerance 100ms)
    merged_final = pd.merge_asof(
        merged_outer, bpm_df, on="timestamp", direction="nearest", tolerance=pd.Timedelta("100ms")
    )

    # # Replace PPG_mV column with BPM values
    # merged_final["PPG_mV"] = merged_final["BPM"]
    # merged_final.drop(columns=["BPM"], inplace=True)  # Remove redundant BPM column

    # --- Addition Start: Interpolation & Highlighting ---
    # Save a mask of where missing values originally exist in the merged dataset
    missing_mask = merged_final.isna()
    
    # Define columns to fill:
    # EEG channels (columns starting with "ch" or in ["X", "Y", "Z"])
    eeg_cols = [col for col in merged_final.columns if col.startswith("ch") or col in ["X", "Y", "Z"]]
    # GSR columns
    gsr_cols = ["GSR_range", "Skin_conductance_uS", "Skin_resistence_kOhms"]
    # BPM column (replacing PPG_mV)
    bpm_col = "BPM"
    ppg_col = "PPG_mV"
    # Interpolate EEG channels, PPG and BPM using linear interpolation
    if eeg_cols:
        merged_final[eeg_cols] = merged_final[eeg_cols].interpolate(method="linear")
    merged_final[bpm_col] = merged_final[bpm_col].interpolate(method="linear")
    merged_final[ppg_col] = merged_final[ppg_col].interpolate(method="linear")
    
    # For GSR values, use forward-fill and then back-fill for any remaining gaps
    merged_final[gsr_cols] = merged_final[gsr_cols].fillna(method="ffill")
    merged_final.fillna(method="bfill", inplace=True)
    # --- Addition End ---
    
    # Save final dataset to Excel with highlighting
    subj_name = Path.split("\\")[-1]
    output_excel = os.path.join(Path, rf"merged_highlighted_{subj_name}.xlsx")  # --- Addition: Excel output path
    merged_final.to_excel(output_excel, index=False)
    
    # --- Addition: Highlighting imputed cells in Excel ---
    wb = load_workbook(output_excel)
    ws = wb.active
    highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    cols = merged_final.columns.tolist()
    # Data starts from row 2 in Excel (row 1 is header)
    for r_idx, row in enumerate(merged_final.itertuples(index=False), start=2):
        for c_idx, col in enumerate(cols, start=1):
            # If the original cell was missing, highlight it
            if missing_mask.loc[merged_final.index[r_idx-2], col]:
                ws.cell(row=r_idx, column=c_idx).fill = highlight_fill
                
    wb.save(output_excel)
    print(f"Merged dataset with interpolated and highlighted cells saved as '{output_excel}'.")
    # Also save CSV if needed
    output_csv = "merged_clean.csv"
    merged_final.to_csv(os.path.join(Path, output_csv), index=False)
    print(f"Merged dataset CSV saved as '{output_csv}'.")
    # --- End Highlighting Addition ---

def main():
    
    Path = "data/thesis_study/subject_003"
    eeg_path = r""
    gsr_path = r""
    
    for f in os.listdir(Path):
        if f.endswith("easy"):
            eeg_path = os.path.join(Path, f.replace("easy", "tsv"))
            os.rename(os.path.join(Path, f), eeg_path)
        if f.endswith("csv"):
            gsr_path = os.path.join(Path, f)
    merge_device_datasets(eeg_path, gsr_path, Path)

if __name__ == "__main__":
    main()
