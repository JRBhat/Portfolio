import pandas as pd
import numpy as np
import scipy.signal as signal
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import os
from tests.diagnostics_heart_rate import diagnose_ppg


def calculate_bpm_from_ppg(df):
    """Process the PPG signal and compute BPM using peak detection and a rolling window approach."""
    # diagnose_ppg(df)
    
    # ————— FIXED TIME CONVERSION —————
    # Shimmer timestamps are in ms since epoch → parse accordingly
    raw_ts = pd.to_numeric(df["timestamp"], errors="coerce")                  # ← FIXED
    df["timestamp"] = pd.to_datetime(raw_ts, unit="ms")                       # ← FIXED
    
    base_timestamp = df["timestamp"].iloc[0]
    time = (df["timestamp"] - base_timestamp).dt.total_seconds()               # ← FIXED
    # ————————————————————————————————————
    
    # Extract PPG
    ppg_signal = df["PPG_mV"]
    
    # Sampling rate
    fs = 1.0 / np.mean(np.diff(time))
    print("Sampling frequency (fs):", fs)
    
    # Bandpass filter 0.5–5 Hz
    lowcut, highcut = 0.5, 5.0
    nyq = fs / 2.0
    if highcut >= nyq:
        highcut = nyq * 0.99
    if lowcut <= 0:
        lowcut = nyq * 0.01
    low  = lowcut / nyq
    high = highcut / nyq
    b, a = signal.butter(3, [low, high], btype='band')
    ppg_filtered = signal.filtfilt(b, a, ppg_signal)
    
    # Smooth (0.5 s MA)
    window_size = max(1, int(fs * 0.5))
    ppg_smoothed = np.convolve(ppg_filtered,
                               np.ones(window_size) / window_size,
                               mode='same')
    
    # ————— FIXED PEAK DETECTION THRESHOLD —————
    # Lower from 75th to 50th percentile so fewer beats are missed
    threshold = np.percentile(ppg_smoothed, 50)                              # ← FIXED
    peaks, _ = signal.find_peaks(ppg_smoothed,
                                 height=threshold,
                                 distance=int(fs * 0.5))
    if len(peaks) == 0:
        print("No peaks detected in PPG signal.")
        return pd.DataFrame(columns=["timestamp", "BPM"])
    peak_times = time[peaks]
    # ————————————————————————————————————
    
    # ————— FIXED BPM CALCULATION (COUNT-BASED) —————
    window_length = 10  # seconds
    bpm_values = []
    bpm_times  = []
    
    for t in peak_times:
        recent = peak_times[(peak_times > t - window_length) & (peak_times <= t)]
        count = len(recent)
        if count > 1:
            bpm = count * (60.0 / window_length)                               # ← FIXED
            bpm_values.append(bpm)
            bpm_times.append(t)
    # ————————————————————————————————————
    
    # Convert back to timestamps
    bpm_timestamps = [base_timestamp + pd.to_timedelta(t, unit="s") 
                      for t in bpm_times]
    
    return pd.DataFrame({
        "timestamp": bpm_timestamps,
        "BPM":       bpm_values
    })



def merge_device_datasets(d1path, d2path, Path):
    """Merge EEG and GSR datasets, replacing PPG_mV with BPM values, fill missing values, and highlight the imputed cells."""
    
    # Load EEG dataset (expected to contain timepoint and EEG values)
    d1 = pd.read_csv(d1path, sep="\t") 
    
    # Load GSR dataset, adjusting headers
    d2 = pd.read_csv(d2path, header=None, skiprows=3, sep="\t", usecols=[0,1,2,3,4])  

    # Set proper column headers
    d1.columns = ['ch1', 'ch2', 'ch3', 'ch4', 'ch5', 'ch6', 'ch7', 'ch8', 'X', 'Y', 'Z', 'marker', 'timestamp']
    new_headers = ['timestamp', 'GSR_range', 'Skin_conductance_uS', 'Skin_resistence_kOhms', 'PPG_mV']
    d2.columns = new_headers
    
    # Convert timestamps from Unix milliseconds to datetime for merging purposes
    d1["timestamp"] = pd.to_datetime(d1["timestamp"], unit="ms")
    d2["timestamp"] = pd.to_datetime(d2["timestamp"], unit="ms")
    
    # Sort data by timestamp
    d1 = d1.sort_values("timestamp")
    d2 = d2.sort_values("timestamp")
    
    # Create a copy of d2 for BPM calculation with numeric timestamps (in ms)
    d2_for_bpm = d2.copy()
    # Convert d2_for_bpm timestamps back to numeric (ms) for BPM calculation
    d2_for_bpm["timestamp"] = d2_for_bpm["timestamp"].astype(np.int64) // 10**6  # --- Correction: integer division to get ms
    
    # Sort d2_for_bpm by timestamp (numeric)
    d2_for_bpm = d2_for_bpm.sort_values("timestamp")
    
    # Compute BPM from PPG signal using the copy with numeric timestamps
    bpm_df = calculate_bpm_from_ppg(d2_for_bpm)

    # Merge EEG and GSR datasets using an outer join
    merged_outer = pd.merge(d1, d2, on="timestamp", how="outer", sort=True)

    # Merge BPM results, replacing PPG_mV with BPM using a nearest match (tolerance 100ms)
    merged_final = pd.merge_asof(
        merged_outer, bpm_df, on="timestamp", direction="nearest", tolerance=pd.Timedelta("100ms")
    )

    # # Replace PPG_mV column with BPM values
    # merged_final["PPG_mV"] = merged_final["BPM"]
    # merged_final.drop(columns=["BPM"], inplace=True)  # Remove redundant BPM column

    # --- Addition Start: Interpolation & Highlighting ---
    # Save a mask of where missing values originally exist in the merged dataset
    missing_mask = merged_final.isna()
    
    # Define columns to fill:
    # EEG channels (columns starting with "ch" or in ["X", "Y", "Z"])
    eeg_cols = [col for col in merged_final.columns if col.startswith("ch") or col in ["X", "Y", "Z"]]
    # GSR columns
    gsr_cols = ["GSR_range", "Skin_conductance_uS", "Skin_resistence_kOhms"]
    # BPM column (replacing PPG_mV)
    bpm_col = "BPM"
    ppg_col = "PPG_mV"
    # Interpolate EEG channels, PPG and BPM using linear interpolation
    if eeg_cols:
        merged_final[eeg_cols] = merged_final[eeg_cols].interpolate(method="linear")
    merged_final[bpm_col] = merged_final[bpm_col].interpolate(method="linear")
    merged_final[ppg_col] = merged_final[ppg_col].interpolate(method="linear")
    
    # For GSR values, use forward-fill and then back-fill for any remaining gaps
    merged_final[gsr_cols] = merged_final[gsr_cols].fillna(method="ffill")
    merged_final.fillna(method="bfill", inplace=True)
    # --- Addition End ---
    
    # Save final dataset to Excel with highlighting
    subj_name = Path.split("\\")[-1]
    output_excel = os.path.join(Path, rf"merged_highlighted_{subj_name}_bpmcorrected.xlsx")  # --- Addition: Excel output path
    merged_final.to_excel(output_excel, index=False)
    
    # --- Addition: Highlighting imputed cells in Excel ---
    wb = load_workbook(output_excel)
    ws = wb.active
    highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    cols = merged_final.columns.tolist()
    # Data starts from row 2 in Excel (row 1 is header)
    for r_idx, row in enumerate(merged_final.itertuples(index=False), start=2):
        for c_idx, col in enumerate(cols, start=1):
            # If the original cell was missing, highlight it
            if missing_mask.loc[merged_final.index[r_idx-2], col]:
                ws.cell(row=r_idx, column=c_idx).fill = highlight_fill
                
    wb.save(output_excel)
    print(f"Merged dataset with interpolated and highlighted cells saved as '{output_excel}'.")
    # Also save CSV if needed
    output_csv = "merged_clean.csv"
    merged_final.to_csv(os.path.join(Path, output_csv), index=False)
    print(f"Merged dataset CSV saved as '{output_csv}'.")
    # --- End Highlighting Addition ---
    return output_excel
