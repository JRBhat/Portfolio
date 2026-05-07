"""
EEG and GSR dataset merge pipeline.

Input formats:
- EEG: tab-separated ``.tsv`` / ``.easy`` file with 13 columns
  (ch1–ch8, X, Y, Z, marker, timestamp).  Timestamps are UNIX milliseconds.
- GSR: Shimmer Consensys CSV in either English (tab-separated, dot decimal)
  or German (semicolon-separated, comma decimal) export format.  Timestamps
  are UNIX milliseconds.

Output:
- A single ``.xlsx`` workbook placed inside the subject directory, with the
  EEG and GSR streams outer-joined on a shared datetime timestamp column.
- BPM is computed from the PPG channel via
  :func:`calculate_bpm_best_practices.calculate_bpm_from_ppg` and merged in
  with a ±100 ms tolerance.
- Missing values are filled (linear interpolation for EEG/BPM; forward-fill
  for GSR columns) and the imputed cells are highlighted yellow
  (``FFFF00``) in the output workbook.
"""
import logging
import pandas as pd
import numpy as np
import scipy.signal as signal
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import os

# from calculate_bpm import calculate_bpm_from_ppg
from calculate_bpm_best_practices import calculate_bpm_from_ppg

logger = logging.getLogger(__name__)

# CEST offset applied when converting UNIX-ms timestamps to local datetime.
LOCAL_TZ_OFFSET = pd.Timedelta(hours=2)

# EEG column names, in order, as produced by the NIC .easy/.tsv export.
EEG_COLUMNS = ['ch1', 'ch2', 'ch3', 'ch4', 'ch5', 'ch6', 'ch7', 'ch8',
               'X', 'Y', 'Z', 'marker', 'timestamp']

# GSR columns that are forward-filled rather than interpolated.
GSR_FORWARD_FILL_COLUMNS = ["GSR_range", "Skin_conductance_uS", "Skin_resistence_kOhms"]

# Hex fill colour used to highlight imputed (originally missing) cells.
IMPUTED_CELL_HIGHLIGHT_COLOR = "FFFF00"

# Tolerance window for the BPM merge_asof join.
BPM_MERGE_TOLERANCE = pd.Timedelta("100ms")


def load_shimmer_csv_auto(filepath):
    """
    Load a Shimmer Consensys GSR CSV file, automatically detecting locale/export format.
    Handles:
      - 'sep=\t' (tab, dot-decimal, English)
      - ';' separator (comma-decimal, German)
    Returns: Clean DataFrame with consistent column names.
    """
    # --- Detect format from first few lines ---
    with open(filepath, 'r', encoding='utf-8') as f:
        head = [next(f) for _ in range(4)]

    # Check for "sep=\t" declaration
    if any("sep=" in line for line in head):
        sep = '\t'
        decimal = '.'
        skiprows = 1  # skip the "sep=\t" line
    else:
        sep = ';'
        decimal = ','
        skiprows = 0

    # --- Read file header and data ---
    df = pd.read_csv(filepath, sep=sep, decimal=decimal, skiprows=3, usecols=[0,1,2,3,4], header=None)
    df.columns = ['timestamp', 'GSR_range', 'Skin_conductance_uS', 'Skin_resistence_kOhms', 'PPG_mV']

    # Convert timestamp column to datetime
    df["timestamp"] = pd.to_datetime(df["timestamp"], unit="ms") + LOCAL_TZ_OFFSET

    # Sort data
    df = df.sort_values("timestamp").reset_index(drop=True)
    return df


def merge_device_datasets(d1path, d2path, Path):
    """Merge EEG and GSR datasets, replacing PPG_mV with BPM values, fill missing values, and highlight the imputed cells."""

    # Get subject name from path and define output Excel path
    subj_name = os.path.basename(Path)
    output_excel = os.path.join(Path, rf"merged_highlighted_{subj_name}_bpmcorrected.xlsx")

    # --- Check if output file already exists ---
    if not os.path.exists(output_excel):
        # --- Load EEG dataset ---
        d1 = pd.read_csv(d1path, sep="\t")
        d1.columns = EEG_COLUMNS
        d1["timestamp"] = pd.to_datetime(d1["timestamp"], unit="ms") + LOCAL_TZ_OFFSET
        d1 = d1.sort_values("timestamp")


        # --- Load GSR dataset (auto-detecting format) ---
        d2 = load_shimmer_csv_auto(d2path)

        # --- Prepare for BPM calculation ---
        d2_for_bpm = d2.copy()
        d2_for_bpm["timestamp"] = d2_for_bpm["timestamp"].astype("int64") // 10**6  # ms
        d2_for_bpm = d2_for_bpm.sort_values("timestamp")

        # --- Compute BPM from PPG signal ---
        bpm_df = calculate_bpm_from_ppg(d2_for_bpm)

        # --- Merge EEG and GSR datasets ---
        merged_outer = pd.merge(d1, d2, on="timestamp", how="outer", sort=True)

        # --- Merge BPM results, replacing PPG_mV with BPM (nearest 100ms) ---
        merged_final = pd.merge_asof(
            merged_outer, bpm_df, on="timestamp", direction="nearest", tolerance=BPM_MERGE_TOLERANCE
        )

        # --- Handle missing values with interpolation and forward/backward fill ---

        missing_mask = merged_final.isna()

        eeg_cols = [c for c in merged_final.columns if c.startswith("ch") or c in ["X","Y","Z"]]
        bpm_col = "BPM"
        ppg_col = "PPG_mV"

        if eeg_cols:
            merged_final[eeg_cols] = merged_final[eeg_cols].interpolate(method="linear")
        if bpm_col in merged_final:
            merged_final[bpm_col] = merged_final[bpm_col].interpolate(method="linear")
        if ppg_col in merged_final:
            merged_final[ppg_col] = merged_final[ppg_col].interpolate(method="linear")

        merged_final[GSR_FORWARD_FILL_COLUMNS] = merged_final[GSR_FORWARD_FILL_COLUMNS].ffill()
        merged_final.bfill(inplace=True)

        # --- Save final dataset to Excel ---
        merged_final.to_excel(output_excel, index=False)

        # --- Highlight imputed cells ---
        wb = load_workbook(output_excel)
        ws = wb.active
        highlight_fill = PatternFill(
            start_color=IMPUTED_CELL_HIGHLIGHT_COLOR,
            end_color=IMPUTED_CELL_HIGHLIGHT_COLOR,
            fill_type="solid",
        )

        # --- Apply highlights based on missing_mask ---
        cols = merged_final.columns.tolist()
        for r_idx, _ in enumerate(merged_final.itertuples(index=False), start=2):
            for c_idx, col in enumerate(cols, start=1):
                if missing_mask.loc[merged_final.index[r_idx-2], col]:
                    ws.cell(row=r_idx, column=c_idx).fill = highlight_fill

        wb.save(output_excel)
        logger.info("Merged dataset with interpolated and highlighted cells saved as '%s'.", output_excel)

        return output_excel

    else:
        logger.info("%s already exists. Skipping merge and proceeding to FFA calculation...", output_excel)
        return output_excel


def main():
    pass


if __name__ == "__main__":
    main()
