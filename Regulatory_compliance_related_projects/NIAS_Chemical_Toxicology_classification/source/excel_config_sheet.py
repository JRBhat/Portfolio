# excel_utils.py
import os
from datetime import datetime
import configparser
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def write_config_sheet(self, missing_file_path: str):
    """
    Same signature & logic — writes 'CONFIG' sheet in provided Excel file.
    Returns True on success, False otherwise.
    """
    try:
        if not missing_file_path:
            return False
        missing_file_path = os.path.expanduser(missing_file_path)
        if not os.path.exists(missing_file_path):
            self.log_field.append(f"⚠️ write_config_sheet: file does not exist: {missing_file_path}")
            return False

        wb = load_workbook(missing_file_path)
        if "CONFIG" in wb.sheetnames:
            del wb["CONFIG"]
        ws = wb.create_sheet("CONFIG")

        ws.append(["Generated on", datetime.now().isoformat()])
        ws.append([])

        ws.append(["GUI INPUTS"])
        gui_entries = [
            ("config_path", self.config_path_field.text().strip()),
            ("xlsx_path", self.xlsx_path_field.text().strip()),
            ("db_path", self.db_path_field.text().strip()),
            ("template_path", self.template_path_field.text().strip()),
            ("out_path", self.out_path_field.text().strip()),
            ("missing_path", self.missing_path_field.text().strip()),
        ]
        for k, v in gui_entries:
            ws.append([k, v])

        ws.append([])

        cfg_path = self.config_path_field.text().strip()
        ws.append(["CONFIG FILE CONTENTS"])
        if cfg_path and os.path.exists(cfg_path):
            try:
                cp = configparser.ConfigParser()
                cp.read(cfg_path, encoding='utf-8')
                for section in cp.sections():
                    ws.append([f"[{section}]"])
                    for key, val in cp.items(section):
                        ws.append([key, val])
                    ws.append([])
            except Exception as e:
                ws.append([f"Error reading config file: {e}"])
        else:
            ws.append(["No config file found or path empty."])

        for col in ws.columns:
            maxlen = 0
            for cell in col:
                if cell.value:
                    l = len(str(cell.value))
                    if l > maxlen:
                        maxlen = l
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(10, min(80, maxlen + 2))

        wb.save(missing_file_path)

        self.log_field.append(f"ℹ️ CONFIG sheet written to {missing_file_path}")
        return True
    except Exception as e:
        try:
            self.log_field.append(f"⚠️ Failed to add CONFIG sheet: {e}")
        except Exception:
            pass
        return False
