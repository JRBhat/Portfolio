#!/usr/bin/env python3
"""
merge_nias_restore_maxrow_append.py

VSCode-friendly script to merge user Excel files' "(FOR_USER)READY_TO_MERGE_DB" sheets
into a DB workbook (NIAS_Substance and NIAS_CL_Type) and produce a CSV report.

Behavior:
 - PRESERVES CAS wrappers (exact string used; only trims leading/trailing whitespace).
 - Exact-match existence check in NIAS_Substance. If found -> highlight that row (light blue).
 - Newly appended NIAS_Substance rows are highlighted in dark green.
 - Appends to each sheet use ws.max_row + 1 (the workbook's max_row) — i.e. "append at the end".
 - Produces CSV report with one row per processed input row describing the actions.
 - DRY_RUN option: when enabled, the DB workbook is NOT saved (but the report is still written).

Edit DB_PATH, USER_FILES, OUT_PATH and REPORT_PATH at the top as needed.
"""

from pathlib import Path
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
import traceback

# ---------------------------
# CONFIGURE PATHS FOR DEBUG
# ---------------------------


OUT_PATH = None     # if None auto-generated next to DB_PATH
REPORT_PATH = None  # if None auto-generated next to OUT_PATH
DRY_RUN = False      # if True: do not save the DB workbook (report still written)

USER_SHEET_NAME = "(FOR_USER)READY_TO_MERGE_DB"
DEBUG = True  # set to False to disable debug prints

# ---------------------------
# Helpers for file selection
# ---------------------------
def find_latest_user_file(folder: Path, prefix: str = "missing_", suffix: str = ".xlsx") -> Path | None:
    """Return the latest file in folder matching pattern like missing_YYYYMMDD_HHMMSS.xlsx."""
    files = [f for f in folder.glob(f"*{suffix}") if f.name.startswith(prefix)]
    if not files:
        return None

    def extract_dt_from_name(path: Path):
        try:
            # Example: missing_20250821_165154.xlsx
            base = path.stem  # "missing_20250821_165154"
            parts = base.split("_")
            if len(parts) < 3:
                return None
            date_str, time_str = parts[1], parts[2]
            return datetime.strptime(date_str + time_str, "%Y%m%d%H%M%S")
        except Exception:
            return None

    files_with_dt = [(f, extract_dt_from_name(f)) for f in files]
    files_with_dt = [(f, dt) for f, dt in files_with_dt if dt is not None]
    if not files_with_dt:
        return None

    # Pick latest by datetime
    latest_file, _ = max(files_with_dt, key=lambda x: x[1])
    return latest_file

# ---------------------------
# Column mappings
# ---------------------------
MAP_SUBSTANCE = {
    "CAS_KEY":    "CAS",
    "PUB_SMILES": "SMILES",
    "IUPAC_DE":   "IUPAC_DE",
    "IUPAC_EN":   "IUPAC_EN",
    "Cramer":     "Cramer",
    "FCM":        "FCM",
    "FT_EN":      "FT_EN",
    "FT_DE":      "FT_DE",
    "FT_HZ":      "FT_HZ",
}

MAP_CL_TYPE = {
    "CAS_KEY":     "CAS",
    "IUPAC_DE":    "IUPAC_DE",
    "CL_DE":       "CL_DE",
    "CL_EN":       "CL_EN",
    "TypeFlag_CL": "TypeFlag_CL",
}

# Fills
EXISTING_ROW_FILL = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")  # light blue
ADDED_ROW_FILL    = PatternFill(start_color="006400", end_color="006400", fill_type="solid")  # dark green

# ---------------------------
# Helpers
# ---------------------------
def as_string_preserve_wrappers(value):
    """Return string with wrappers preserved (only strip leading/trailing whitespace).
       Return None for None/NaN/empty."""
    if value is None:
        return None
    try:
        if pd.isna(value):
            return None
    except Exception:
        pass
    s = str(value).strip()
    return s if s != "" else None

def ensure_sheet_with_headers(wb, sheet_name, headers):
    """Ensure sheet exists and contains at least the provided headers (append missing headers)."""
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        existing_headers = [(cell.value or "") for cell in next(ws.iter_rows(max_row=1))]
        missing = [h for h in headers if h not in existing_headers]
        if missing:
            current_cols = len(existing_headers)
            for i, mh in enumerate(missing, start=1):
                ws.cell(row=1, column=current_cols + i, value=mh)
        return ws
    else:
        ws = wb.create_sheet(sheet_name)
        for col_idx, h in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=h)
        return ws

def header_list_from_sheet(ws):
    """Return list of header names (row 1) for ws."""
    return [(cell.value or "") for cell in next(ws.iter_rows(max_row=1))]

def write_row_using_max_row(ws, row_values):
    """
    Write row_values to ws at row index ws.max_row + 1 (append at sheet's max_row end).
    Returns the row number written.
    """
    target = ws.max_row + 1
    # If sheet is empty, ws.max_row is 0 or 1 — writing at 1 is okay; headers may be overwritten if misused.
    for col_idx, val in enumerate(row_values, start=1):
        ws.cell(row=target, column=col_idx, value=val)
    if DEBUG:
        print(f"write_row_using_max_row -> wrote to row {target} (ws.max_row was {ws.max_row - 0})")
    return target

def fill_entire_row(ws, rownum, fill):
    """Apply the given PatternFill to every cell in the row up to ws.max_column."""
    for col in range(1, ws.max_column + 1):
        ws.cell(row=rownum, column=col).fill = fill

def build_cas_lookup_from_sheet(ws, cas_header_name="CAS"):
    """Return mapping cas_string -> row_number from ws (Exact string as stored in sheet)."""
    headers = header_list_from_sheet(ws)
    if cas_header_name not in headers:
        return {}, None
    cas_col_idx = headers.index(cas_header_name) + 1
    cas_to_row = {}
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        cell = row[cas_col_idx - 1]
        if cell.value is None:
            continue
        cas_val = as_string_preserve_wrappers(cell.value)
        if cas_val:
            cas_to_row[cas_val] = cell.row
    return cas_to_row, cas_col_idx

def find_missing_mapped_fields_for_sheet(row, mapping, sheet_label):
    """
    Given a dataframe row (pandas Series), and a mapping dict (user_col -> target_col),
    return a list of (user_col, target_col) that are considered missing (empty / NaN / whitespace).
    Excludes 'CAS_KEY' from checks here (the caller handles CAS separately).
    """
    missing = []
    skipped_columns = ["CAS_KEY", "FCM", "FT_EN", "FT_DE", "FT_HZ", "CL_DE", "CL_EN"]
    for user_col, target_col in mapping.items():
        if user_col in skipped_columns:
            continue
        raw_val = row.get(user_col, None)
        # Use same emptiness semantics as as_string_preserve_wrappers: None or pd.NaN or empty -> missing
        v = as_string_preserve_wrappers(raw_val)
        if v is None:
            missing.append((user_col, target_col))
    return missing


# ---------------------------
# Core merge + report (append via max_row)
# ---------------------------


def merge_user_files_into_db_and_report(db_path: Path, user_files: list[Path],
                                        out_path: Path, report_path: Path, dry_run: bool=False):
    """
    Merge using append-by-max_row for both sheets. Writes DB to out_path (unless dry_run=True)
    and writes CSV report to report_path.

    Strict validation: Before copying/appending for a given row, ensure ALL mapped user columns
    (except CAS_KEY) have values for the sheet(s) that would be written. If any required value
    is missing, print a warning listing missing fields and abort the merge (no DB saved).
    """
    # Load or create DB workbook
    if db_path.exists():
        wb = load_workbook(db_path)
        if DEBUG: print(f"Loaded DB workbook from {db_path}")
    else:
        wb = Workbook()
        if "Sheet" in wb.sheetnames and wb["Sheet"].cell(1,1).value is None:
            del wb["Sheet"]
        if DEBUG: print("Created new DB workbook (no existing file).")

    # Determine headers (prefer existing)
    if "NIAS_Substance" in wb.sheetnames:
        sub_headers = header_list_from_sheet(wb["NIAS_Substance"])
    else:
        sub_headers = [v for v in MAP_SUBSTANCE.values()]

    if "NIAS_CL_Type" in wb.sheetnames:
        cl_headers = header_list_from_sheet(wb["NIAS_CL_Type"])
    else:
        cl_headers = [v for v in MAP_CL_TYPE.values()]

    # Ensure sheets exist with headers
    ws_sub = ensure_sheet_with_headers(wb, "NIAS_Substance", sub_headers)
    ws_cl  = ensure_sheet_with_headers(wb, "NIAS_CL_Type", cl_headers)

    # Build lookup maps using exact stored values
    cas_to_row, _ = build_cas_lookup_from_sheet(ws_sub, cas_header_name="CAS")
    cl_cas_to_row, _ = build_cas_lookup_from_sheet(ws_cl, cas_header_name="CAS")
    cl_cas_set = set(cl_cas_to_row.keys())

    report_rows = []
    ts_print = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    appended_sub_count = 0
    appended_cl_count = 0
    highlighted_count = 0
    processed_rows = 0

    # We'll abort on the first validation error and provide details.
    for user_file in user_files:
        if DEBUG: print(f"\nProcessing user file: {user_file}")
        if not user_file.exists():
            if DEBUG: print("  -> file not found, skipping.")
            continue

        try:
            df = pd.read_excel(user_file, sheet_name=USER_SHEET_NAME, dtype=object)
        except Exception as e:
            if DEBUG: print(f"  -> could not read sheet '{USER_SHEET_NAME}' from {user_file}: {e}")
            continue

        df.columns = [str(c).strip() for c in df.columns]
        if "CAS_KEY" not in df.columns:
            if DEBUG: print("  -> sheet missing CAS_KEY, skipping file.")
            continue

        for idx, row in df.iterrows():
            processed_rows += 1
            excel_row = idx + 2
            raw_cas = row.get("CAS_KEY", None)
            cas_value = as_string_preserve_wrappers(raw_cas)

            entry = {
                "timestamp": ts_print,
                "source_file": str(user_file),
                "excel_row": excel_row,
                "cas_key": cas_value,
                "action_substance": None,
                "substance_row": None,
                "action_cl": None,
                "cl_row": None,
                "notes": ""
            }

            if cas_value is None:
                entry["action_substance"] = "skipped"
                entry["action_cl"] = "skipped"
                entry["notes"] = "empty CAS_KEY"
                report_rows.append(entry)
                continue

            # --- VALIDATION LOGIC ---
            # If CAS already exists: we will only append CL if it's missing in CL sheet.
            # If CAS does not exist: we will append Substance AND CL (both must be validated before writing).
            # For whichever sheet(s) are to be written, ensure all mapped user columns (except CAS_KEY) are present.

            # Prepare lists of required checks
            will_write_substance = cas_value not in cas_to_row
            will_write_cl = False
            if cas_value in cas_to_row:
                # CAS exists in substance: append CL only if CL doesn't exist for this CAS
                if cas_value not in cl_cas_set:
                    will_write_cl = True
            else:
                # new CAS -> we will append CL as well
                will_write_cl = True

            # Validate substance fields if we will write substance
            if will_write_substance:
                missing_sub = find_missing_mapped_fields_for_sheet(row, MAP_SUBSTANCE, "NIAS_Substance")
            else:
                missing_sub = []

            # Validate CL fields if we will write CL
            if will_write_cl:
                missing_cl = find_missing_mapped_fields_for_sheet(row, MAP_CL_TYPE, "NIAS_CL_Type")
            else:
                missing_cl = []

            # If any missing -> print detailed warning and abort entire merge (stop copying)
            if missing_sub or missing_cl:
                # Build readable message
                msgs = []
                if missing_sub:
                    msgs.append(
                        "NIAS_Substance missing mapped input(s): " +
                        ", ".join([f"{u} -> {t}" for (u, t) in missing_sub])
                    )
                if missing_cl:
                    msgs.append(
                        "NIAS_CL_Type missing mapped input(s): " +
                        ", ".join([f"{u} -> {t}" for (u, t) in missing_cl])
                    )
                full_msg = (
                    "\nERROR: Missing required user input detected. Aborting copy.\n"
                    f" Source file: {user_file}\n"
                    f" Excel row: {excel_row}\n"
                    f" CAS_KEY: {cas_value}\n"
                    + "\n".join(msgs)
                    + "\n\nPlease fill the missing cells in the user's sheet and re-run."
                )
                # Print + raise to abort (caller will catch)
                print(full_msg)
                raise RuntimeError(full_msg)

            # --- end validation, proceed with previous logic ---

            # If substance exists, highlight and possibly append CL
            if cas_value in cas_to_row:
                sub_rownum = cas_to_row[cas_value]
                # highlight existing substance row (light blue)
                fill_entire_row(ws_sub, sub_rownum, EXISTING_ROW_FILL)
                highlighted_count += 1
                entry["action_substance"] = "existing"
                entry["substance_row"] = sub_rownum

                # Ensure CL exists for this CAS; if not append to CL using max_row
                if cas_value in cl_cas_set:
                    entry["action_cl"] = "already_exists"
                    entry["cl_row"] = cl_cas_to_row[cas_value]
                else:
                    cl_headers_current = header_list_from_sheet(ws_cl)
                    cl_row_values = []
                    for hdr in cl_headers_current:
                        val_to_write = None
                        for user_col, target_col in MAP_CL_TYPE.items():
                            if target_col == hdr:
                                if target_col == "CAS":
                                    val_to_write = cas_value
                                else:
                                    v = row.get(user_col, None)
                                    try:
                                        val_to_write = None if pd.isna(v) else v
                                    except Exception:
                                        val_to_write = v
                                break
                        cl_row_values.append(val_to_write)
                    new_cl_rownum = write_row_using_max_row(ws_cl, cl_row_values)
                    appended_cl_count += 1
                    entry["action_cl"] = "appended"
                    entry["cl_row"] = new_cl_rownum
                    cl_cas_set.add(cas_value)
                    cl_cas_to_row[cas_value] = new_cl_rownum

                report_rows.append(entry)
                continue

            # CAS not present in NIAS_Substance -> append substance using max_row
            sub_headers_current = header_list_from_sheet(ws_sub)
            sub_row_values = []
            for hdr in sub_headers_current:
                val_to_write = None
                for user_col, target_col in MAP_SUBSTANCE.items():
                    if target_col == hdr:
                        if target_col == "CAS":
                            val_to_write = cas_value
                        else:
                            v = row.get(user_col, None)
                            try:
                                val_to_write = None if pd.isna(v) else v
                            except Exception:
                                val_to_write = v
                        break
                sub_row_values.append(val_to_write)
            new_sub_rownum = write_row_using_max_row(ws_sub, sub_row_values)
            appended_sub_count += 1
            cas_to_row[cas_value] = new_sub_rownum

            # Highlight the newly added substance row (dark green)
            fill_entire_row(ws_sub, new_sub_rownum, ADDED_ROW_FILL)

            entry["action_substance"] = "appended"
            entry["substance_row"] = new_sub_rownum

            # Append CL row using max_row (independent)
            cl_headers_current = header_list_from_sheet(ws_cl)
            cl_row_values = []
            for hdr in cl_headers_current:
                val_to_write = None
                for user_col, target_col in MAP_CL_TYPE.items():
                    if target_col == hdr:
                        if target_col == "CAS":
                            val_to_write = cas_value
                        else:
                            v = row.get(user_col, None)
                            try:
                                val_to_write = None if pd.isna(v) else v
                            except Exception:
                                val_to_write = v
                        break
                cl_row_values.append(val_to_write)
            new_cl_rownum = write_row_using_max_row(ws_cl, cl_row_values)
            # highlight the newly added CL row
            fill_entire_row(ws_cl, new_cl_rownum, ADDED_ROW_FILL)

            appended_cl_count += 1
            entry["action_cl"] = "appended"
            entry["cl_row"] = new_cl_rownum
            cl_cas_set.add(cas_value)
            cl_cas_to_row[cas_value] = new_cl_rownum

            report_rows.append(entry)

    # Prepare save paths
    ts_fname = datetime.now().strftime("%Y%m%d_%H%M%S")
    if report_path is None:
        report_path = db_path.with_name(f"{db_path.stem}_report_v{ts_fname}.csv")

    # Ensure parent folders exist
    db_parent = db_path.parent
    db_parent.mkdir(parents=True, exist_ok=True)
    report_parent = report_path.parent
    report_parent.mkdir(parents=True, exist_ok=True)

    # Save DB workbook unless dry-run
    if dry_run:
        print(f"\nDRY RUN: DB workbook not saved (would have been: {db_path})")
    else:
        wb.save(db_path)
        print(f"\nSaved merged DB to: {db_path}")

    # Save CSV report
    df_report = pd.DataFrame(report_rows, columns=[
        "timestamp", "source_file", "excel_row", "cas_key",
        "action_substance", "substance_row", "action_cl", "cl_row", "notes"
    ])
    df_report.to_csv(report_path, index=False, encoding="utf-8-sig")
    print(f"Saved CSV report to: {report_path}")

    summary = {
        "timestamp": ts_print,
        "processed_rows": processed_rows,
        "highlighted": highlighted_count,
        "appended_sub": appended_sub_count,
        "appended_cl": appended_cl_count,
        "out_path": str(db_path) if not dry_run else f"(dry-run) {db_path}",
        "report_path": str(report_path)
    }
    return summary

# ---------------------------
# Run (for debug in VS Code)
# ---------------------------
if __name__ == "__main__":
    # Update these paths to point to your local NIAS database and user files for testing.
    DB_PATH = Path(r"path\to\NIAS_TEST.xlsx")
    USER_FOLDER = Path(r"path\to\user_files_folder")


    try:
        ts_fname = datetime.now().strftime("%Y%m%d_%H%M%S")

        latest_file = find_latest_user_file(USER_FOLDER)
        if latest_file is None:
            print(f"No suitable user files found in {USER_FOLDER}")
            exit(1)

        USER_FILES = [latest_file]  # overwrite dynamically
        print(f"Using latest user file: {latest_file}")

        if OUT_PATH:
            out_path = Path(OUT_PATH)
        else:
            out_path = DB_PATH.with_name(f"{DB_PATH.stem}_DB_v{ts_fname}.xlsx")

        if REPORT_PATH:
            report_path = Path(REPORT_PATH)
        else:
            report_path = out_path.with_name(f"{DB_PATH.stem}_report_v{ts_fname}.csv")

        print("DB_PATH:", DB_PATH)
        print("USER_FOLDER:", USER_FOLDER)
        print("Selected USER_FILES:", USER_FILES)
        print("OUT_PATH:", out_path)
        print("REPORT_PATH:", report_path)
        print("DRY_RUN:", DRY_RUN)

        summary = merge_user_files_into_db_and_report(DB_PATH, USER_FILES, out_path, report_path, dry_run=DRY_RUN)

        print("\n=== Summary ===")
        for k, v in summary.items():
            print(f"{k}: {v}")

    except RuntimeError as exc:
        print("\nMerge aborted due to validation error.")
    except Exception as exc:
        print("Unexpected error:")
        traceback.print_exc()
