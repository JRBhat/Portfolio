"""
reporting.py
============
Contains the :class:`Reporting` QThread that orchestrates the complete NIAS
reporting pipeline in a background thread so the GUI remains responsive.

Pipeline stages (executed sequentially inside :meth:`Reporting.run`)
---------------------------------------------------------------------
1. **Config loading** (:meth:`read_config`) – reads the INI config file for
   paths, ISD list, and pass/fail label text.  Parameters that can be set
   via the GUI (uncertainty factors, specimen mass, CAS column flag) are
   intentionally *not* read from the config to avoid overwriting user input.

2. **Database parsing** (:func:`~parsing.parse_db`) – loads all substance
   records from the NIAS DB Excel file into ``self.db_substances``.

3. **Excel parsing** (:meth:`parse_xlsx`) – reads the laboratory "Auswertung"
   sheet and categorises each row into ``self.substances`` (identified, with
   CAS), ``self.unidentified`` (no CAS, no name), or ``self.alkan`` (alkane).

4. **Missing-substance handling**
   (:func:`~missing_substances_and_bad_cas_handling.save_new_substances_separately`) –
   identifies substances absent from the DB, queries PubChem for enrichment,
   and writes a debug Excel file.

5. **Merging** (:func:`~merging.merge_db_and_excel_substances`) – applies DB
   or PubChem data to every substance object in place.

6. **Export** (:func:`~exports.export_document`) – generates the final Word
   report with all six tables populated.

Threading model
---------------
:class:`Reporting` inherits from :class:`~PySide6.QtCore.QThread`.  The GUI
connects to the ``text_changed`` signal to receive log messages, and to the
``finished`` signal to detect completion and enable/disable controls.

``self.success`` and ``self.last_exception`` are set at the end of
:meth:`run` so the GUI can display a success or error dialog.

Template filename conventions
------------------------------
The language (``_de_`` / ``_en_``) and template type
(``lebensmittel`` / ``kosmetikum``) are detected by inspecting substrings of
the template filename.  A missing tag raises
:class:`InformationMissingInFilenameError` and defaults to ``DE`` / ``L``.
"""

import os
import logging
import pandas as pd
import configparser
from openpyxl import load_workbook
import locale

from PySide6.QtCore import QThread, Signal

import warnings
warnings.simplefilter('ignore', DeprecationWarning)
import urllib3
warnings.simplefilter('ignore', urllib3.exceptions.InsecureRequestWarning)

from exports import export_document
from parsing import parse_db
from substance import Substance
from merging import merge_db_and_excel_substances
from missing_substances_and_bad_cas_handling import save_new_substances_separately
from internal_imports import cur_ft_dict, cramer_TTC_mapping, mapping_fcm_group


class InformationMissingInFilenameError(Exception):
    """Exception raised when a valid template class or language is not specified in the filename of the template"""
    def __init__(self, message):
        super().__init__(message)
        self.message = message

class Reporting(QThread):
    text_changed = Signal(str)

    def __init__(self):
        super().__init__()

        # existing attributes...
        self.substances = []
        self.unidentified = []
        self.alkan = []
        self.db_substances = {}
        self.xlsx_path = ''
        self.outname = ''
        self.toxtree_path = ''
        self.toxtree_jar = ''
        self.java_path = ''
        self.db_path = ''
        self.legacy_db_path = ''
        self.plastchem_db_path = ''
        self.rebuild_db = False

        self.template_path = ''
        self.template = ''

        # These will now be set from GUI instead of config
        self.uf_tenax = 1
        self.uf_real = 1
        self.uf_exposition = 1
        self.K_gewicht = 60
        self.cas_column = False
        
        self.fts = {}
        self.cur_ft = 'a'
        self.language = ''
        self.template_type = ''
        self.passed_yes = ''
        self.passed_no = ''
        self.passed_maybe = ''
        self.running = True
        self.config_path = None
        self.debug_path = ''

        # NEW: outcome tracking
        self.success = False
        self.last_exception = None

    def stop(self):
        self.running = False

    def run(self):
        # Start by assuming failure — only set to True on full success
        self.success = False
        self.last_exception = None

        try:
            # Logs the software version and emits it via text_changed to the GUI
            self.text_changed.emit('NIAS Reporting (version 21.11.24)')

            # Check if the config path is set
            if not self.config_path:
                self.text_changed.emit("Error: Config path not set.")
                self.success = False
                return

            # initialize settings or configurations from config file.
            try:
                self.read_config(self.config_path)
                self.text_changed.emit("Config file loaded successfully.")
            except Exception as e:
                self.text_changed.emit(f"Error reading config: {e}")
                self.last_exception = str(e)
                self.success = False
                return

            # Basic validation that required paths are present
            missing = []
            if not self.db_path:
                missing.append("db_path")
            if not self.xlsx_path:
                missing.append("xlsx_path")
            if not self.out_path:
                missing.append("out_path")
            if not self.template:
                missing.append("template")
            if missing:
                msg = f"Error: Missing required paths: {', '.join(missing)}"
                self.text_changed.emit(msg)
                self.last_exception = msg
                self.success = False
                return

            # Calls parse_db() to load substances from the primary database - NIAS DB
            self.text_changed.emit(f'Reading NIAS DB {self.db_path}')
            df_eu, df_eu_grp = parse_db(self.db_path, self.db_substances, self.language, self.template_type)

            # Calls parse_xlsx() to process the user-selected file
            self.text_changed.emit(f'Reading {self.xlsx_path}')
            self.text_changed.emit('=======' + '=' * (len(self.xlsx_path) + 1))
            _ = self.parse_xlsx(self.xlsx_path, mode='list')

            saved_dict_for_merging = save_new_substances_separately(
                self.substances, self.db_substances, self.isd,
                self.db_path, self.debug_path, self.text_changed
            )

            merge_db_and_excel_substances(self.substances, self.db_substances, saved_dict_for_merging, self.isd)

            export_document(
                self.xlsx_path, self.out_path, self.text_changed, self.template_path,
                substances=self.substances,
                isd=self.isd,
                language=self.language,
                passed_yes=self.passed_yes,
                passed_maybe=self.passed_maybe,
                passed_no=self.passed_no,
                uf_real=self.uf_real,
                uf_exposition=self.uf_exposition,
                alkan=self.alkan,
                cas_column=self.cas_column,
                unidentified=self.unidentified,
                db_path=self.db_path,
                cramer_TTC_mapping=cramer_TTC_mapping,
                cur_ft_dict=cur_ft_dict,
                mapping_fcm_group=mapping_fcm_group,
                df_eu=df_eu,
                df_eu_grp=df_eu_grp,
                K_gewicht=self.K_gewicht,
                template_type=self.template_type
            )
            self.text_changed.emit("Export completed successfully.")

            # If we reached this point without exception, mark success
            self.success = True
            self.last_exception = None

        except Exception as e:
            # Log exception and notify GUI; keep the exception message for later inspection
            logging.exception("Error in Reporting thread")
            self.last_exception = str(e)
            self.text_changed.emit(f"An unexpected error occurred. Check logs. ({e})")
            self.success = False
        finally:
            # run() returns, QThread will emit finished()
            return

    def read_config(self, config_path):
        """
        Reads configuration values from the specified config file and sets up
        the necessary paths and settings for reporting.

        Only sets attributes from the config if the corresponding attribute
        is not already set (GUI selections take precedence).
        
        NOTE: Parameters like K_gewicht, uf_real, uf_exposition, uf_tenax, and cas_column
        are now set directly from the GUI and will NOT be read from the config file.
        """
        self.text_changed.emit(f"Reading config {config_path}...")
        if not os.path.exists(config_path):
            raise FileNotFoundError(f"Config file not found: {config_path}")

        config = configparser.ConfigParser()
        config.read(config_path)

        # Input section: prefer GUI values, otherwise take from config if present.
        try:
            if 'Input' in config:
                input_sec = config['Input']
                if not self.xlsx_path and 'xlsxPath' in input_sec:
                    self.xlsx_path = input_sec.get('xlsxPath', fallback='').strip()
                if not self.db_path and 'dbPath' in input_sec:
                    self.db_path = input_sec.get('dbPath', fallback='').strip()
        except Exception as e:
            logging.exception("Error parsing Input section")
            raise

        # Output section
        try:
            if 'Output' in config:
                output_sec = config['Output']
                if not getattr(self, 'out_path', '') and 'outPath' in output_sec:
                    self.out_path = output_sec.get('outPath', fallback='').strip()
                if not getattr(self, 'debug_path', '') and 'missingPath' in output_sec:
                    self.debug_path = output_sec.get('missingPath', fallback='').strip()
        except Exception as e:
            logging.exception("Error parsing Output section")
            raise

        # Report section
        try:
            if 'Report' in config:
                report_sec = config['Report']
                # template only if not already provided by GUI
                if not self.template and 'Template' in report_sec:
                    raw_template = report_sec.get('Template', fallback='').strip()
                    if raw_template:
                        # store template filename without extension
                        self.template = os.path.splitext(os.path.basename(raw_template))[0]
                        # also keep full path in template_path if it looks like a path
                        if os.path.exists(raw_template):
                            self.template_path = raw_template

                if not getattr(self, 'isd', None) and 'ISD' in report_sec:
                    # ISD is expected to be a list literal in the ini
                    self.isd = eval(report_sec.get('ISD', '[]'))

                # NOTE: Parameters below are now set from GUI, so we skip reading them from config
                # They are kept here as comments for reference
                # if 'Umrechnungsfaktor_Tenax' in report_sec:
                #     try:
                #         self.uf_tenax = eval(report_sec.get('Umrechnungsfaktor_Tenax', '1'))
                #     except Exception:
                #         self.uf_tenax = 1
                # if 'Umrechnungsfaktor_Real' in report_sec:
                #     try:
                #         self.uf_real = eval(report_sec.get('Umrechnungsfaktor_Real', '1').replace(',', '.'))
                #     except Exception:
                #         self.uf_real = 1
                # if 'Umrechnungsfaktor_Exposition' in report_sec:
                #     try:
                #         self.uf_exposition = eval(report_sec.get('Umrechnungsfaktor_Exposition', '1'))
                #     except Exception:
                #         self.uf_exposition = 1
                # if 'CASColumn_Table3' in report_sec:
                #     try:
                #         self.cas_column = eval(report_sec.get('CASColumn_Table3', 'False'))
                #     except Exception:
                #         self.cas_column = False
                # if 'K_gewicht' in report_sec:
                #     try:
                #         self.K_gewicht = eval(report_sec.get('K_gewicht', '60'))
                #     except Exception:
                #         self.K_gewicht = 60
        except Exception as e:
            logging.exception("Error parsing Report section")
            raise

        # Text section: passed texts must be present for the determined language
        try:
            # determine language from template if not already set
            if self.template and not self.language:
                try:
                    if '_de_' in self.template.lower():
                        self.language = 'DE'
                        # Avoid forcing locale on user's system if not available - optional:
                        try:
                            locale.setlocale(locale.LC_ALL, 'de_DE')
                        except Exception:
                            pass
                    elif '_en_' in self.template.lower():
                        self.language = 'EN'
                    else:
                        raise InformationMissingInFilenameError(f"Language not found in filename: {self.template.lower()}")
                except InformationMissingInFilenameError as e:
                    logging.error(f"{e}, defaulting to DE.")
                    self.language = 'DE'
                else:
                    self.text_changed.emit(f"Language {self.language} detected from template.")

            # Template type
            try:
                if self.template and not self.template_type:
                    if "lebensmittel" in self.template.lower():
                        self.template_type = 'L'
                    elif "kosmetikum" in self.template.lower():
                        self.template_type = 'K'
                    else:
                        raise InformationMissingInFilenameError(f"Template type not provided in the filename: {self.template.lower()}")
            except InformationMissingInFilenameError as e:
                logging.error(f"{e}, defaulting to L")
                self.template_type = 'L'
            else:
                if self.template:
                    self.text_changed.emit(f"Classification type {self.template_type} set successfully")

            if 'Text' in config and self.language:
                text_sec = config['Text']
                # safe access to keys like PassedYes_DE
                self.passed_yes = text_sec.get(f'PassedYes_{self.language}', fallback=self.passed_yes)
                self.passed_no = text_sec.get(f'PassedNo_{self.language}', fallback=self.passed_no)
                self.passed_maybe = text_sec.get(f'PassedMaybe_{self.language}', fallback=self.passed_maybe)
        except Exception:
            logging.exception("Error parsing Text section")
            raise

        self.text_changed.emit(f" ✅ Config file {config_path} read successfully.")
        if self.template:
            self.text_changed.emit(f"✅ Template {self.template} ready.")

    def parse_xlsx(self, xlsx_path, mode):
        """
        Parses an Excel file containing substance data.

        Args:
            xlsx_path (str): Path to the Excel file.
            mode (str): Parsing mode, either 'list' to append to `self.substances`
                        or 'dict' to return a dictionary.

        Returns:
            dict: Parsed substances (if `mode` is 'dict').
        """

        substances = {}
        if not os.path.exists(xlsx_path):
            self.text_changed.emit(f"❌ Error: Excel file {xlsx_path} not found.")
            return substances

        self.text_changed.emit(f"Reading Excel file {xlsx_path}...")
        wb = load_workbook(filename=xlsx_path, data_only=True)

        auswertung = None
        for sheetname in wb.sheetnames:
            if sheetname.strip() == 'Auswertung':
                auswertung = wb[sheetname]
                break

        if auswertung is None:
            raise ValueError('Worksheet Auswertung not found')

        # New xlsx file format
        if auswertung['A1'].value == 'Retention Time FID [min]':
            data = auswertung.values
            columns = next(data)
            df = pd.DataFrame(data, columns=columns)

            for ind, row in df.iterrows():
                s = Substance()
                select_hit = row.get('Select Hit') if hasattr(row, 'get') else row['Select Hit']

                # only selects the data with a 'x' marked
                if select_hit == 'x':
                    rt = row['Retention Time FID [min]']
                    s.rt = rt
                    s.name = row.get('Substance')
                    s.cas = row.get('CAS')
                    s.qual = row.get('Quality')
                    try:
                        s.quant = row['Concentration (as DEHP) \nmg/L Probe']
                    except Exception:
                        # attempt fallback if heading differs slightly
                        try:
                            for colname in row.index:
                                if "Concentration" in str(colname) and "DEHP" in str(colname):
                                    s.quant = row[colname]
                                    break
                        except Exception:
                            s.quant = None

                    if s.cas:
                        if mode == 'list':
                            self.substances.append(s)
                        elif mode == 'dict':
                            substances[rt] = s
                    else:
                        self.unidentified.append(s)
        else:  # old format
            row = 19
            done = False

            while row < auswertung.max_row:
                s = Substance()
                rt = auswertung[f'A{row}'].value

                if isinstance(rt, float):
                    s.rt = rt
                    s.area = auswertung[f'C{row}'].value
                    s.name = auswertung[f'D{row}'].value
                    s.cas = auswertung[f'E{row}'].value
                    s.qual = auswertung[f'F{row}'].value
                    s.quant = auswertung[f'H{row}'].value

                    if s.cas is not None:
                        if mode == 'list':
                            self.substances.append(s)
                        elif mode == 'dict':
                            substances[rt] = s
                    elif s.name == 'Alkan':
                        self.alkan.append(s)
                    elif s.area != 'Blindwert' and (s.name is None or s.name == 'cyc'):
                        self.unidentified.append(s)

                    if rt is None:
                        done = True

                row += 1

        if self.substances:
            self.text_changed.emit(f'{len(self.substances)} substances with CAS read')
        else:
            self.text_changed.emit(f'{len(substances)} substances with CAS read')

        self.text_changed.emit("✅ LAB Excel file read successfully....")
        return substances
