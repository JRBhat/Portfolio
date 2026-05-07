import openpyxl
from openpyxl.styles import PatternFill

def highlight_duplicates(file_path, sheet_name=None):
    """
    Opens the given Excel file and highlights duplicate rows with a yellow fill.
    Only duplicate occurrences (rows after the first instance) are highlighted.
    The function saves the updated workbook as "highlighted_duplicates.xlsx".

    Args:
        file_path (str): Path to the input Excel file.
        sheet_name (str, optional): Name of the sheet to process; if None, uses the active sheet.

    Returns:
        tuple: (output_file, duplicate_row_indices) where output_file is the path to the highlighted file,
               and duplicate_row_indices is a set of row numbers that were marked as duplicates.
    """
    # Load the workbook and select the sheet (default: active)
    wb = openpyxl.load_workbook(file_path)
    ws = wb[sheet_name] if sheet_name else wb.active

    # Define yellow fill for highlighting duplicate rows
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    seen_rows = {}            # Dictionary to record first occurrence: {row_values_tuple: row_number}
    duplicate_row_indices = set()  # Set to hold row indices of duplicates

    # Iterate over all rows (assumes data starts at row 1; adjust min_row if you have headers)
    for row in ws.iter_rows(min_row=1, values_only=False):
        # Create a tuple of cell values for the current row
        # row_values = tuple(cell.value for cell in row)
        row_values = tuple(cell.value for cell in row[:-1])# ignores last column
        if row_values in seen_rows:
            # Only highlight subsequent occurrences (duplicates), not the first occurrence
            duplicate_row_indices.add(row[0].row)
        else:
            seen_rows[row_values] = row[0].row

    # Apply the yellow fill to every cell in each duplicate row
    for row_index in duplicate_row_indices:
        for cell in ws[row_index]:
            cell.fill = yellow_fill

    output_file = "highlighted_duplicates.xlsx"
    wb.save(output_file)
    print(f"Highlighted duplicate rows saved in '{output_file}'.")
    return output_file, duplicate_row_indices

def delete_highlighted_rows(file_path, sheet_name=None):
    """
    Opens the given Excel file and deletes any rows whose first cell is highlighted in yellow.
    Assumes that duplicate rows were previously highlighted with a yellow fill.
    The function saves the updated workbook as "duplicates_removed.xlsx".

    Args:
        file_path (str): Path to the Excel file with highlighted rows.
        sheet_name (str, optional): Name of the sheet to process; if None, uses the active sheet.
    """
    wb = openpyxl.load_workbook(file_path)
    ws = wb[sheet_name] if sheet_name else wb.active

    rows_to_delete = []
    # Check each row: if the first cell's fill color indicates yellow, mark that row for deletion.
    for row in ws.iter_rows(min_row=1):
        # Depending on how openpyxl stores colors, the RGB value might include a leading "00"
        fill_color = row[0].fill.start_color.rgb
        if fill_color in ["FFFF00", "00FFFF00"]:
            rows_to_delete.append(row[0].row)

    # Delete rows in reverse order to avoid index shifting issues
    for row_index in sorted(rows_to_delete, reverse=True):
        ws.delete_rows(row_index)

    cleaned_file = "duplicates_removed.xlsx"
    wb.save(cleaned_file)
    print(f"Deleted highlighted rows. Updated file saved as '{cleaned_file}'.")

if __name__ == "__main__":
    # Update this path to point to your local NIAS database Excel file.
    input_file = r"path\to\NIAS_DB_FILLED_updated.xlsx"

    # Step 1: Highlight duplicate rows
    highlighted_file, duplicate_rows = highlight_duplicates(input_file)
    print("Duplicate row indices:", duplicate_rows)

    # Step 2: Delete the highlighted (duplicate) rows
    delete_highlighted_rows(highlighted_file)
