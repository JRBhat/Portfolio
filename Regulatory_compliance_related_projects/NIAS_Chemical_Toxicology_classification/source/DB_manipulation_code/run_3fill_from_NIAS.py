import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Load input files
x_df = pd.read_excel(os.environ.get("BADCAS_FILLED_PATH", "data/ZDB_main/ALLBADCAS_merged_CAS_filled.xlsx"))
_nias_db_path = os.environ.get("NIAS_DB_PATH", "data/ZDB_main/NIAS_ZDB.xlsx")
nias_sub_df = pd.read_excel(_nias_db_path, sheet_name="NIAS_Substance")
nias_cl_df = pd.read_excel(_nias_db_path, sheet_name="NIAS_CL_Type")

# Standardize CAS columns
x_df['CAS_KEY'] = x_df['CAS_KEY'].astype(str).str.strip().str.replace('#', '', regex=False)
nias_sub_df['CAS'] = nias_sub_df['CAS'].astype(str).str.strip()
nias_cl_df['CAS'] = nias_cl_df['CAS'].astype(str).str.strip()

# Fill missing values from NIAS_Substance and NIAS_CL_Type
highlight_rows = []

columns_map_sub = {
    'IUPAC_DE': 'IUPAC_DE',
    'IUPAC_EN': 'IUPAC_EN',
    'SMILES': 'PUB_SMILES',
    'Cramer': 'Cramer',
    'FCM': 'FCM',
    'FT_EN': 'FT_EN',
    'FT_DE': 'FT_DE',
    'FT_HZ': 'FT_HZ'
}
columns_map_cl = {
    'CL_DE': 'CL_DE',
    'CL_EN': 'CL_EN',
    'TypeFlag_CL': 'TypeFlag_CL'
}

# Iterate over rows
for idx, row in x_df.iterrows():
    cas = row['CAS_KEY']
    if cas in nias_sub_df['CAS'].values:
        sub_data = nias_sub_df[nias_sub_df['CAS'] == cas].iloc[0]
        cl_data = nias_cl_df[nias_cl_df['CAS'] == cas].iloc[0] if cas in nias_cl_df['CAS'].values else None
        updated = False

        for src_col, target_col in columns_map_sub.items():
            if pd.isna(row[target_col]) and not pd.isna(sub_data[src_col]):
                x_df.at[idx, target_col] = sub_data[src_col]
                updated = True

        if cl_data is not None:
            for src_col, target_col in columns_map_cl.items():
                if pd.isna(row[target_col]) and not pd.isna(cl_data[src_col]):
                    x_df.at[idx, target_col] = cl_data[src_col]
                    updated = True

        if updated:
            highlight_rows.append(idx)

# Remove exact duplicates, keeping first
x_df = x_df.drop_duplicates(keep='first')

# Save to Excel
output_file = "X_enriched.xlsx"
x_df.to_excel(output_file, index=False)

# Apply highlighting to modified rows
wb = load_workbook(output_file)
ws = wb.active
fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

for idx in highlight_rows:
    if idx + 2 <= ws.max_row:  # +2 because Excel rows are 1-based and header is row 1
        for col in range(1, ws.max_column + 1):
            ws.cell(row=idx + 2, column=col).fill = fill

wb.save(output_file)
print(f"Enriched file saved as: {output_file}")
