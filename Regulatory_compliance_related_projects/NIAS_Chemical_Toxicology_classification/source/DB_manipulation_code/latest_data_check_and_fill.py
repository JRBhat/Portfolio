
import os
import pandas as pd
import requests
import logging
import re
from deepdiff import DeepDiff
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


# ----------------------------
# Logger Setup
# ----------------------------
formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')

general_logger = logging.getLogger('general_errors')
general_logger.setLevel(logging.ERROR)
gen_file_handler = logging.FileHandler('errors.log', mode='w')
gen_file_handler.setFormatter(formatter)
general_logger.addHandler(gen_file_handler)
gen_console_handler = logging.StreamHandler()
gen_console_handler.setFormatter(formatter)
general_logger.addHandler(gen_console_handler)

diffs_logger = logging.getLogger("diffs_logger")
diffs_logger.setLevel(logging.INFO)
diffs_handler = logging.FileHandler("diffs.log", mode='w')
diffs_handler.setFormatter(formatter)
diffs_logger.addHandler(diffs_handler)
diffs_console_handler = logging.StreamHandler()
diffs_console_handler.setFormatter(formatter)
diffs_logger.addHandler(diffs_console_handler)


# ----------------------------
# API Query Functions
# ----------------------------
import requests

def query_chemspider(identifier):
    """Query ChemSpider for CAS, IUPAC, SMILES, and German name based on input type."""
    user_agent = 'Mozilla/5.0'
    base_url = "https://www.chemspider.com/api/search?value={}"
    identifier = identifier.strip()
    url = base_url.format(identifier)
    
    try:
        response = requests.get(url, headers={"User-Agent": user_agent}, verify=True)
        response.raise_for_status()
        data = response.json()
        
        if "Records" in data and data["Records"]:
            record = data["Records"][0]
            title = record.get("Title")
            cas = record.get("Cas")
            smiles = record.get("StructuralIdentifiers", {}).get("Smiles")
            iupac = record.get("IUPACName")

            # Extract German name
            german_name = None
            synonyms = record.get("Synonyms", [])
            for synonym in synonyms:
                if synonym.get("Language") == "de":
                    german_name = synonym.get("Name")
                    break  # Stop at the first German name found

            return {"Identifier": identifier, "Title": title,"CAS": cas, "SMILES": smiles, "IUPAC": iupac, "German Name": german_name}
        else:
            general_logger.error(f"ChemSpider: No records found for identifier: {identifier}")
    except requests.RequestException as e:
        general_logger.error(f"ChemSpider query failed for identifier {identifier}: {e}")
    return None

def query_pubchem(identifier, query_type="cas"):
    """Query PubChem using an identifier (CAS, IUPAC, or SMILES) to retrieve
    canonical properties and CAS number (from synonyms).
    """
    # Clean the identifier string
    identifier = identifier.strip()
    
    # Determine which endpoint to use
    if query_type.lower() in ["cas", "iupac", "name"]:
        input_type = "name"
    elif query_type.lower() == "smiles":
        input_type = "smiles"
    else:
        raise ValueError("Invalid query_type. Use 'cas', 'iupac', or 'smiles'.")
    
    # URL to fetch canonical properties
    prop_url = ("https://pubchem.ncbi.nlm.nih.gov/rest/pug/compound/{}/{}/"
                "property/CanonicalSMILES,IUPACName,MolecularFormula/JSON").format(input_type, identifier)
    
    try:
        response = requests.get(prop_url, verify=True)
        response.raise_for_status()
        data = response.json()
        
        properties = data.get("PropertyTable", {}).get("Properties", [])
        if not properties:
            general_logger.error(f"PubChem: No properties found for identifier: {identifier}")
            return None
        
        result = {
            "Identifier": identifier,
            "SMILES": properties[0].get("CanonicalSMILES"),
            "IUPAC": properties[0].get("IUPACName"),
            "MolecularFormula": properties[0].get("MolecularFormula"),
            "CAS": None  # Will update after checking synonyms
        }
        
        # Query synonyms to try and extract the CAS number
        syn_url = ("https://pubchem.ncbi.nlm.nih.gov/rest/pug/compound/{}/{}/synonyms/JSON").format(input_type, identifier)
        syn_response = requests.get(syn_url, verify=True)
        syn_response.raise_for_status()
        syn_data = syn_response.json()
        
        synonyms = syn_data.get("InformationList", {}).get("Information", [{}])[0].get("Synonym", [])
        cas_pattern = re.compile(r'^\d{2,7}-\d{2}-\d$')
        for syn in synonyms:
            if cas_pattern.match(syn):
                result["CAS"] = syn
                break
        
        return result
    except requests.RequestException as e:
        general_logger.error(f"PubChem query failed for identifier {identifier}: {e}")
    
    return None

# ----------------------------
# Validation and Update Functions
# ----------------------------

# CAS-validation-function
def validate_cas(cas):
    cas_pattern = re.compile(r'^\d{2,7}-\d{2}-\d$')
    if cas_pattern.match(cas):
        positions = np.arange(9, -1, -1)  # Array of decreasing positions
        ziffern = np.array(list(f"{int(''.join(cas.split('-'))):010d}"), dtype=int)  # Convert to int array
        betrag = positions[:-1] * ziffern[:-1]  # Element-wise multiplication
        return int(np.sum(betrag) % 10) == ziffern[-1]  # Validate check digit
    return False  # Return False if CAS format is invalid


def capitalize_first_alpha(s):
    return next((s[:i] + c.upper() + s[i+1:] for i, c in enumerate(s) if c.isalpha()), s)


def update_missing_values(identifier, idtype, row_dict):
    
    data_pubchem = query_pubchem(identifier, idtype)
    if data_pubchem:
        if data_pubchem.get("CAS"):
            row_dict["CAS"] = data_pubchem.get("CAS") if row_dict['CAS'] != data_pubchem.get("CAS") else row_dict['CAS']
        if data_pubchem.get("SMILES"):
            row_dict["SMILES"] = data_pubchem.get("SMILES") if row_dict['SMILES'] != data_pubchem.get("SMILES") else row_dict['SMILES']
        if data_pubchem.get("IUPAC"):
            row_dict["IUPAC"] = capitalize_first_alpha(data_pubchem.get("IUPAC")) if row_dict['IUPAC'] != capitalize_first_alpha(data_pubchem.get("IUPAC")) else row_dict['IUPAC']
            
    else: # try getting cas with smiles code if iupac not available
        general_logger.error(f"PubChem: Failed to retrieve data for identifier: {identifier}")

    data_chemspider = query_chemspider(identifier)
    
    if data_chemspider:
        if data_chemspider.get("German Name"):
            row_dict["IUPAC_DE"] = data_chemspider.get("German Name") if row_dict['IUPAC_DE'] != data_chemspider.get("German Name") else row_dict['IUPAC_DE']    
    else: # try getting cas with smiles code if iupac not available
        general_logger.error(f"ChemSpider: Failed to retrieve data for identifier: {identifier}")
    return row_dict

# ----------------------------
# Highlighting Function
# ----------------------------
def highlight_error_rows(file_path, invalid_indices, error_indices, output_file):
    """
    Highlight rows in the Excel file.
      - Rows with invalid CAS (BAD CAS) are highlighted in red.
      - Rows with query or fallback errors are highlighted in yellow.
    """
    wb = load_workbook(file_path)
    ws = wb.active  

    red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")

    # Highlight rows with invalid CAS in red.
    for idx in set(invalid_indices):
        excel_row = idx + 2  # DataFrame index to Excel row (header offset)
        for col in range(1, ws.max_column + 1):
            ws.cell(row=excel_row, column=col).fill = red_fill

    # Highlight rows with query/fallback errors in yellow (if not already highlighted red).
    for idx in set(error_indices) - set(invalid_indices):
        excel_row = idx + 2
        for col in range(1, ws.max_column + 1):
            ws.cell(row=excel_row, column=col).fill = yellow_fill

    wb.save(output_file)
    print(f"Rows with errors highlighted and saved as '{output_file}'")

# ----------------------------
# Main Processing Function
# ----------------------------
def main():
    # read excel file as df
    # Load the Excel file (adjust the filename if needed)
    df = pd.read_excel(os.environ.get("NIAS_DB_PATH", "data/NIAS_ZDB.xlsx"))
    
    invalid_indices = []  # Rows with invalid (BAD) CAS numbers.
    error_indices = []    # Rows where queries or fallbacks generated errors.


    # loop through all rows and first check if cas numbers are correct
    for idx, row in df.iterrows():
        
        row_dict = {
        "CAS" : row["CAS"],
        "IUPAC" : row["IUPAC_EN"],
        "IUPAC_DE" : row["IUPAC_DE"],
        "SMILES" : row["SMILES"],  
        }
        
        old_dict = row_dict.copy()
        # check if cas is nan or "-"
        if not pd.isna(row_dict["CAS"]) and row_dict["CAS"] != "-" and row_dict["CAS"] != "":
            
            
            # # check if cas is invalid highlight its row RED and skip this row
            if not validate_cas(row_dict["CAS"]):# if invalid cas
                invalid_indices.append(idx)
                continue # Skip further processing for this row.
            
            
            # check if valid iupac available, if yes then check cas using pubchem
            if not pd.isna(row_dict["IUPAC"]) and row_dict["IUPAC"] not in ["-", ""]: 
                row_dict = update_missing_values(row_dict["IUPAC"], "iupac", row_dict)
            
            # if iupac missing then check cas using available smile using pubchem  
            elif not pd.isna(row_dict["SMILES"]) and row_dict["SMILES"] not in ["-", ""]:
                row_dict = update_missing_values(row_dict["SMILES"], "smiles", row_dict)

            else:
                general_logger.error(f"Both smiles and iupac missing for {row_dict}, trying with CAS")
                row_dict = update_missing_values(row_dict["CAS"], "cas", row_dict)
                if row_dict == old_dict:
                    general_logger.error(f"Failed after fallback with {row_dict["CAS"]}. \n\t Unchanged : {row_dict}")
                    error_indices.append(idx)
        
        # When CAS is missing.
        else:
            
            # check if valid iupac available, if yes then check cas using pubchem
            if not pd.isna(row_dict["IUPAC"]) and row_dict["IUPAC"] not in ["-", ""]: 
                row_dict = update_missing_values(row_dict["IUPAC"], "iupac", row_dict)
            
            # if iupac missing then check cas using available smile using pubchem  
            elif not pd.isna(row_dict["SMILES"]) and row_dict["SMILES"] not in ["-", ""]:
                row_dict = update_missing_values(row_dict["SMILES"], "smiles", row_dict)

            else:
                general_logger.error(f"Both smiles and iupac missing for {row_dict}, skipping due to no CAS")
                error_indices.append(idx)
        
        # Log and update if any values have changed.        
        if old_dict != row_dict:
            diff = DeepDiff(old_dict, row_dict)
            diffs_logger.info(diff)
            df.at[idx, "IUPAC_EN"] = row_dict["IUPAC"]
            df.at[idx, "IUPAC_DE"] = row_dict["IUPAC_DE"]
            df.at[idx, "SMILES"] = row_dict["SMILES"]
            df.at[idx, "CAS"] = row_dict["CAS"]
    
    updated_file = "updated_file.xlsx"
    df.to_excel(updated_file, index=False)
    print(f"Update complete. New file saved as '{updated_file}'.")

    # Highlight rows with invalid CAS and/or query errors.
    highlight_error_rows(updated_file, invalid_indices, error_indices, "highlighted_output.xlsx")
    print("Please check 'errors.log', 'diffs.log', and the highlighted Excel file for details.")


def highlight_invalid_rows(file_path, output_file, invalid_indices):
    """Highlight rows with invalid CAS numbers in red."""
    wb = load_workbook(file_path)
    ws = wb.active  

    red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")

    for idx in invalid_indices:
        excel_row = idx + 2  # Convert DataFrame index to Excel row (1-based index, skipping header)
        for col in range(1, ws.max_column + 1):  # Highlight the entire row
            ws.cell(row=excel_row, column=col).fill = red_fill

    wb.save(output_file)
    print(f"Invalid CAS rows highlighted and saved as '{output_file}'")




if __name__ == "__main__":
    main()