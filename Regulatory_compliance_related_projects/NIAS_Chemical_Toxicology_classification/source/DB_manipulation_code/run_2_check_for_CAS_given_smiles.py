import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
import requests
import re
import os
import logging

# Setup basic logger
logging.basicConfig(level=logging.INFO)
general_logger = logging.getLogger(__name__)

def query_pubchem(identifier, query_type="iupac"):
    """Query PubChem using an IUPAC name and return SMILES and CAS."""
    identifier = identifier.strip()

    if query_type.lower() in ["cas", "iupac", "name"]:
        input_type = "name"
    elif query_type.lower() == "smiles":
        input_type = "smiles"
    else:
        raise ValueError("Invalid query_type. Use 'cas', 'iupac', or 'name'.")

    cid_url = f"https://pubchem.ncbi.nlm.nih.gov/rest/pug/compound/{input_type}/{requests.utils.quote(identifier)}/cids/JSON"
    try:
        cid_response = requests.get(cid_url, timeout=10)
        cid_response.raise_for_status()
        cid_data = cid_response.json()
        cids = cid_data.get("IdentifierList", {}).get("CID", [])
        if not cids:
            general_logger.warning(f"No CID found for identifier: {identifier}")
            return None
        cid = cids[0]

        # Fetch properties
        prop_url = f"https://pubchem.ncbi.nlm.nih.gov/rest/pug/compound/cid/{cid}/property/ConnectivitySMILES,IUPACName,MolecularFormula/JSON"
        prop_response = requests.get(prop_url, timeout=10)
        prop_response.raise_for_status()
        prop_data = prop_response.json()
        props = prop_data.get("PropertyTable", {}).get("Properties", [])
        if not props:
            general_logger.warning(f"No properties found for CID: {cid}")
            return None

        result = {
            "Identifier": identifier,
            "SMILES": props[0].get("ConnectivitySMILES"),
            "IUPAC": props[0].get("IUPACName"),
            "MolecularFormula": props[0].get("MolecularFormula"),
            "CAS": None
        }

        # Fetch synonyms to find CAS
        syn_url = f"https://pubchem.ncbi.nlm.nih.gov/rest/pug/compound/cid/{cid}/synonyms/JSON"
        syn_response = requests.get(syn_url, timeout=10)
        syn_response.raise_for_status()
        syn_data = syn_response.json()

        synonyms = syn_data.get("InformationList", {}).get("Information", [{}])[0].get("Synonym", [])
        cas_pattern = re.compile(r'^\d{2,7}-\d{2}-\d$')
        for syn in synonyms:
            if cas_pattern.match(syn):
                result["CAS"] = syn
                break

        print(f"Found CAS: {result['CAS']} for identifier: {identifier}")
        return result

    except requests.RequestException as e:
        general_logger.error(f"PubChem query failed for identifier {identifier}: {e}")
    except Exception as e:
        general_logger.exception(f"Unexpected error for identifier {identifier}: {e}")

    return None

def update_excel_with_pubchem_cas(file_path):
    df = pd.read_excel(file_path)
    updated_rows = []

    for index, row in df.iterrows():
        if pd.isna(row["PUB_CAS"]):
            cas_result = None
            if pd.notna(row["PUB_SMILES"]):
                cas_result = query_pubchem(row["PUB_SMILES"], query_type="smiles")
            if not cas_result and pd.notna(row["Trivial_Name"]):
                cas_result = query_pubchem(row["Trivial_Name"], query_type="name")

            if cas_result and cas_result["CAS"]:
                df.at[index, "PUB_CAS"] = cas_result["CAS"]
                updated_rows.append(index)

    # Save updated file
    base, ext = os.path.splitext(file_path)
    new_file_path = f"{base}_CAS_filled{ext}"
    df.to_excel(new_file_path, index=False)

    # Reopen with openpyxl to highlight
    wb = openpyxl.load_workbook(new_file_path)
    ws = wb.active

    fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    for row_index in updated_rows:
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row_index + 2, column=col).fill = fill  # +2 accounts for header

    wb.save(new_file_path)
    print(f"Updated file saved as: {new_file_path}")

if __name__ == "__main__":
# Example usage
    update_excel_with_pubchem_cas(os.environ.get("BADCAS_INPUT_PATH", "data/ZDB_main/ALLBADCAS_merged.xlsx"))
