import pandas as pd
import requests
import logging
import re
from deepdiff import DeepDiff
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# ----------------------------
# Logger Setup
# ----------------------------
formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')

general_logger = logging.getLogger('general_errors')
general_logger.setLevel(logging.ERROR)
gen_file_handler = logging.FileHandler('errors.log', mode='w')
gen_file_handler.setFormatter(formatter)
general_logger.addHandler(gen_file_handler)
gen_console_handler = logging.StreamHandler()
gen_console_handler.setFormatter(formatter)
general_logger.addHandler(gen_console_handler)

diffs_logger = logging.getLogger("diffs_logger")
diffs_logger.setLevel(logging.INFO)
diffs_handler = logging.FileHandler("diffs.log", mode='w')
diffs_handler.setFormatter(formatter)
diffs_logger.addHandler(diffs_handler)
diffs_console_handler = logging.StreamHandler()
diffs_console_handler.setFormatter(formatter)
diffs_logger.addHandler(diffs_console_handler)

# ----------------------------
# API Query Functions
# ----------------------------
def query_chemspider(identifier):
    """Query ChemSpider for CAS, IUPAC, SMILES, and German name based on input type."""
    user_agent = 'Mozilla/5.0'
    base_url = "https://www.chemspider.com/api/search?value={}"
    identifier = identifier.strip()
    url = base_url.format(identifier)

    try:
        response = requests.get(url, headers={"User-Agent": user_agent}, verify=True)
        response.raise_for_status()
        data = response.json()
        # TODO: get IUPAC_EN from Chemspider ; get data["Synonyms"]["Name"] if data["Synonyms"]["Language"] == "en" and ACD/IUPAC Name' in data["Synonyms"]["Flags"]
        if "Records" in data and data["Records"]:
            record = data["Records"][0]
            title, cas, iupac = (record.get(k, None) for k in ("Title", "Cas", "IUPACName"))
            smiles = record.get("StructuralIdentifiers", {}).get("Smiles", None)

            synonyms = record.get("Synonyms", [])
            english_name = next((s["Name"] for s in synonyms if s["Language"] == "en" and "ACD/IUPAC Name" in s["Flags"]), None)
            german_name = next((s["Name"] for s in synonyms if s["Language"] == "de"), None)


            return {"Identifier": identifier, "Title": title, "CAS": cas,
                    "SMILES": smiles, "IUPAC": iupac, "German Name": german_name, "English Name": english_name}
        else:
            general_logger.error(f"ChemSpider: No records found for identifier: {identifier}")
    except requests.RequestException as e:
        general_logger.error(f"ChemSpider query failed for identifier {identifier}: {e}")
    return None

def query_pubchem(identifier, query_type="cas"):
    """Query PubChem using an identifier (CAS, IUPAC, or SMILES)."""
    identifier = identifier.strip()
    if query_type.lower() in ["cas", "iupac", "name"]:
        input_type = "name"
    elif query_type.lower() == "smiles":
        input_type = "smiles"
    else:
        raise ValueError("Invalid query_type. Use 'cas', 'iupac', or 'smiles'.")

    prop_url = ("https://pubchem.ncbi.nlm.nih.gov/rest/pug/compound/{}/{}/"
                "property/CanonicalSMILES,IUPACName,MolecularFormula/JSON").format(input_type, identifier)
    try:
        response = requests.get(prop_url, verify=True)
        response.raise_for_status()
        data = response.json()

        properties = data.get("PropertyTable", {}).get("Properties", [])
        if not properties:
            general_logger.error(f"PubChem: No properties found for identifier: {identifier}")
            return None

        result = {
            "Identifier": identifier,
            "SMILES": properties[0].get("CanonicalSMILES"),
            "IUPAC": properties[0].get("IUPACName"),
            "MolecularFormula": properties[0].get("MolecularFormula"),
            "CAS": None
        }

        syn_url = ("https://pubchem.ncbi.nlm.nih.gov/rest/pug/compound/{}/{}/synonyms/JSON").format(input_type, identifier)
        syn_response = requests.get(syn_url, verify=True)
        syn_response.raise_for_status()
        syn_data = syn_response.json()

        synonyms = syn_data.get("InformationList", {}).get("Information", [{}])[0].get("Synonym", [])
        cas_pattern = re.compile(r'^\d{2,7}-\d{2}-\d$')
        for syn in synonyms:
            if cas_pattern.match(syn):
                result["CAS"] = syn
                break

        return result
    except requests.RequestException as e:
        general_logger.error(f"PubChem query failed for identifier {identifier}: {e}")
    return None

# ----------------------------
# Validation and Update Functions
# ----------------------------
def validate_cas(cas):
    """Validate a CAS number using its check digit."""
    cas_pattern = re.compile(r'^\d{2,7}-\d{2}-\d$')
    if cas_pattern.match(cas):
        positions = np.arange(9, -1, -1)
        ziffern = np.array(list(f"{int(''.join(cas.split('-'))):010d}"), dtype=int)
        betrag = positions[:-1] * ziffern[:-1]
        return int(np.sum(betrag) % 10) == ziffern[-1]
    return False

def capitalize_first_alpha(s):
    """Capitalize the first alphabetic character in the string."""
    return next((s[:i] + c.upper() + s[i+1:] for i, c in enumerate(s) if c.isalpha()), s)

def update_missing_values(identifier, idtype, row_dict):
    """Update missing values using PubChem and ChemSpider queries."""
    data_pubchem = query_pubchem(identifier, idtype)
    if data_pubchem:
        if data_pubchem.get("CAS"):
            # Only update if different
            if row_dict["CAS"] != data_pubchem.get("CAS"):
                row_dict["CAS"] = data_pubchem.get("CAS")
        if data_pubchem.get("SMILES"):
            if row_dict["SMILES"] != data_pubchem.get("SMILES"):
                row_dict["SMILES"] = data_pubchem.get("SMILES")
        if data_pubchem.get("IUPAC"):
            new_iupac = capitalize_first_alpha(data_pubchem.get("IUPAC"))
            if row_dict["IUPAC"] != new_iupac:
                row_dict["IUPAC"] = new_iupac
    else:
        general_logger.error(f"PubChem: Failed to retrieve data for identifier: {identifier}")

    data_chemspider = query_chemspider(identifier)
    if data_chemspider:
        if data_chemspider.get("German Name"):
            if row_dict["IUPAC_DE"] != data_chemspider.get("German Name"):
                row_dict["IUPAC_DE"] = data_chemspider.get("German Name")
    else:
        general_logger.error(f"ChemSpider: Failed to retrieve data for identifier: {identifier}")
    return row_dict


def validate_iupacEN(valid_cas, existing_iupacEN, row_dict, row_remarks, error_indices, idx):
    data_chemspider = query_chemspider(valid_cas)

    if data_chemspider:
        chemspider_iupacEN = data_chemspider.get("English Name")
        if chemspider_iupacEN and existing_iupacEN != chemspider_iupacEN:
            row_dict["IUPAC"] = chemspider_iupacEN
            error_msg = f"{existing_iupacEN} changed to {chemspider_iupacEN}"
            row_remarks.append(error_msg)
            general_logger.error(error_msg + f": {row_dict}")
            error_indices.append(idx)
        else:
            error_msg = "IUPAC unchanged"
            row_remarks.append(error_msg)
            general_logger.error(error_msg + f": {row_dict}")
            error_indices.append(idx)

    return row_dict, row_remarks, error_indices, idx

# ----------------------------
# Highlighting Function
# ----------------------------
def highlight_error_rows(file_path, invalid_indices, error_indices, output_file):
    """
    Highlight rows in the Excel file:
      - Invalid CAS rows are highlighted in red.
      - Rows with query/fallback errors are highlighted in yellow.
    """
    wb = load_workbook(file_path)
    ws = wb.active

    red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")

    # Highlight rows with invalid CAS in red.
    for idx in set(invalid_indices):
        excel_row = idx + 2  # Adjust for header offset (Excel rows are 1-indexed)
        for col in range(1, ws.max_column + 1):
            ws.cell(row=excel_row, column=col).fill = red_fill

    # Highlight rows with query/fallback errors in yellow (unless already highlighted red).
    for idx in set(error_indices) - set(invalid_indices):
        excel_row = idx + 2
        for col in range(1, ws.max_column + 1):
            ws.cell(row=excel_row, column=col).fill = yellow_fill

    wb.save(output_file)
    print(f"Rows with errors highlighted and saved as '{output_file}'")

# ----------------------------
# Main Processing Function
# ----------------------------
def main():
    # Update this path to point to your local NIAS database Excel file.
    file_path = r"path\to\NIAS_ZDB_highlighted_Remarks.xlsx"
    df = pd.read_excel(file_path, "Sheet1")

    # Add a new Remarks column for logging error/info messages per row.
    df["Remarks"] = ""

    invalid_indices = []  # Indices for rows with an invalid CAS.
    error_indices = []    # Indices for rows with query/fallback errors.

    for idx, row in df.iterrows():
        row_dict = {
            "CAS": row["CAS"],
            "IUPAC": row["IUPAC_EN"],
            "IUPAC_DE": row["IUPAC_DE"],
            "SMILES": row["SMILES"],
        }
        old_dict = row_dict.copy()
        row_remarks = []  # Collect remarks for this row.

        # If CAS is present
        if not pd.isna(row_dict["CAS"]) and row_dict["CAS"] not in ["-", ""]:
            if not validate_cas(row_dict["CAS"]):
                row_remarks.append("Invalid CAS")
                invalid_indices.append(idx)
                df.at[idx, "Remarks"] = "; ".join(row_remarks)
                continue  # Skip further processing for this row.

            # Update missing values based on available IUPAC or SMILES.
            if not pd.isna(row_dict["IUPAC"]) and row_dict["IUPAC"] not in ["-", ""]:
                row_dict, row_remarks, error_indices, idx = validate_iupacEN(row_dict["CAS"], row_dict["IUPAC"], row_dict, row_remarks, error_indices, idx)
                row_dict = update_missing_values(row_dict["IUPAC"], "iupac", row_dict)
            elif not pd.isna(row_dict["SMILES"]) and row_dict["SMILES"] not in ["-", ""]:
                row_dict = update_missing_values(row_dict["SMILES"], "smiles", row_dict)
            else:
                error_msg = "Both SMILES and IUPAC missing; trying fallback with CAS"
                row_remarks.append(error_msg)
                general_logger.error(f"{error_msg}: {row_dict}")
                row_dict = update_missing_values(row_dict["CAS"], "cas", row_dict)
                if row_dict == old_dict:
                    error_msg = f"Failed after fallback with {row_dict['CAS']}. Row unchanged."
                    row_remarks.append(error_msg)
                    general_logger.error(error_msg)
                    error_indices.append(idx)
        else:
            # When CAS is missing, try updating using IUPAC or SMILES.
            if not pd.isna(row_dict["IUPAC"]) and row_dict["IUPAC"] not in ["-", ""]:
                row_dict = update_missing_values(row_dict["IUPAC"], "iupac", row_dict)
            elif not pd.isna(row_dict["SMILES"]) and row_dict["SMILES"] not in ["-", ""]:
                row_dict = update_missing_values(row_dict["SMILES"], "smiles", row_dict)
            else:
                error_msg = "Both SMILES and IUPAC missing, and no valid CAS"
                row_remarks.append(error_msg)
                general_logger.error(error_msg + f": {row_dict}")
                error_indices.append(idx)

        # If any updates occurred, log the diff and add to remarks.
        if old_dict != row_dict:
            diff = DeepDiff(old_dict, row_dict)
            diffs_logger.info(diff)
            row_remarks.append(f"Updated: {diff}")
            # Update the DataFrame with new values.
            df.at[idx, "IUPAC_EN"] = row_dict["IUPAC"]
            df.at[idx, "IUPAC_DE"] = row_dict["IUPAC_DE"]
            df.at[idx, "SMILES"] = row_dict["SMILES"]
            df.at[idx, "CAS"] = row_dict["CAS"]

        # Save the remarks for this row.
        df.at[idx, "Remarks"] = "; ".join(row_remarks)

    # Save the updated DataFrame (including the Remarks column).
    updated_file = "updated_file_rem.xlsx"
    df.to_excel(updated_file, index=False)
    print(f"Update complete. New file saved as '{updated_file}'.")

    # Highlight the rows with errors in the saved Excel file.
    highlight_error_rows(updated_file, invalid_indices, error_indices, "highlighted_output_rem.xlsx")
    print("Please check 'errors.log', 'diffs.log', and the highlighted Excel file for details.")

if __name__ == "__main__":
    main()
    # cas = "108-94-1"
    # iupac = "Cyclohexanone"
    # smiles = "CC1=C(C2=C(C=C1)C(CC=C2)(C)C)C"
    # print(query_pubchem(smiles, "smiles"))
    # print(query_chemspider(cas))
    # # print(query_pubchem(iupac, "iupac"))
    # # print(query_chemspider(iupac))
