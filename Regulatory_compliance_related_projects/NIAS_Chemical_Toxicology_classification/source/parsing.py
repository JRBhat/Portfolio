"""
parsing.py
==========
Reads the NIAS substance database Excel file and (optionally) legacy Word
document tables, producing the in-memory data structures used throughout the
reporting pipeline.

Primary public functions
------------------------
parse_db(db_path, db_substances, language, template_type)
    Reads multiple sheets from the NIAS DB Excel file and populates the
    ``db_substances`` dictionary.  Also returns the EU2011 and
    Groups_EU2011 DataFrames required for FCM/SML look-ups.

parse_eu_regulation(substances, filename)
    **Legacy** function.  Parses an older EU regulation CSV file and assigns
    FCM/SML values to an already-loaded list of substances.

parse_table(leading_cas_number, t, cols, ft_text, lab_substances)
    Parses a Word document table (``docx.Table``) to extract substance
    metadata.  Retained for backward compatibility with workflows that
    couple a Word file to the Excel results.

Hazard code helpers
-------------------
extend_ft_hazard_code_if_any(df_hazard, ft, language)
    Checks whether a functional-type string is an H-code and, if so,
    appends the human-readable description from the Hcodes sheet.

get_hazard_descp(df_hazard, language, code)
    Looks up a single H-code description in the provided hazard DataFrame.

Language / template type conventions
-------------------------------------
- ``language``: ``'DE'`` (German) or ``'EN'`` (English)
- ``template_type``: ``'L'`` (Lebensmittel / food contact) or
  ``'K'`` (Kosmetikum / cosmetics)
"""

import re
import pandas as pd
from openpyxl import load_workbook
from docx import Document
import math
import numpy as np
from substance import Substance

from utilityFuncs.util import get_text_and_superscript, get_cl_and_inci

def parse_eu_regulation(substances, filename):
    """
    Parses the European Union regulation file to check compliance for substances.

    Args:
        filename (str): Path to the EU regulation CSV file.
    """
    regulation = pd.read_csv(filename, sep='\t')

    for s in substances:
        res = regulation.loc[regulation.CAS == s.cas.lstrip('0')]
        if len(res) >= 1:
            s.fcm = res.FCM.iloc[0]
            s.fcm_name = res.Name.iloc[0]
            sml = res.SML.iloc[0]
            if type(sml) is float and math.isnan(sml):
                sml = '-'
            elif type(sml) is str and ',' in sml:
                sml = float(sml.replace(',', '.'))
            s.sml = str(sml)
            print(f'{s.cas} - Listed ({res.FCM.iloc[0]})')
        else:
            print(f'{s.cas} - Not listed')

def parse_db(db_path, db_substances, language, template_type):
    """
    Parses a database file in Excel format to populate `db_substances`.

    Args:
        filename (str): Path to the Excel database file.
    """
    df = pd.read_excel(db_path, "NIAS_Substance")
    df = df.replace(np.nan, None)
    df_cl = pd.read_excel(db_path, sheet_name="NIAS_CL_Type")
    df_cl = df_cl.replace(np.nan, None)
    df_eu = pd.read_excel(db_path, sheet_name="EU2011")
    df_eu = df_eu.replace(np.nan, None)
    df_eu_grp = pd.read_excel(db_path, sheet_name="Groups_EU2011")
    df_eu_grp = df_eu_grp.replace(np.nan, None)
    df_hazard = pd.read_excel(db_path, "Hcodes")
    df_hazard = df_hazard.replace(np.nan, None)
    cas_numbers = df["CAS"].unique()
    # print(cas_numbers)  # Output unique CAS numbers
    
    if language == "DE":
        if template_type == "L":
            
            grp_indcs = df["FT_DE"].groupby(df["CAS"]).groups
            grp_indcs = {cas: grp_indcs.get(cas, "") for cas in cas_numbers}
        elif template_type == "K":
            grp_indcs = df["FT_DE"].groupby(df["CAS"]).groups
            grp_indcs = {cas: grp_indcs.get(cas, "") for cas in cas_numbers}
        else:
            raise Exception("Template not found")    
        
        #  merge all relevant footnotes (FT_DE) for each CAS number, using the row indices grouped by CAS (stored in grp_indcs)
        fts_combined_for_all_cas = {}
        for cas, indices in grp_indcs.items():
            # Retrieve the actual footnote texts from the DataFrame for each index, filtering out None values
            footnotes = [df.loc[i, "FT_DE"] for i in indices if df.loc[i, "FT_DE"] is not None]
            
            # Removes duplicates from the list of footnotes while preserving their original order.
            # Trick: dict.fromkeys(list) keeps only the first occurrence of each item.
            deduped = list(dict.fromkeys(footnotes))
            
            # If there are any deduplicated footnotes, store them under the current CAS.
            # and If the list is empty (no valid footnotes), store None.
            fts_combined_for_all_cas[cas] = deduped if deduped else None
    

    elif language == "EN":
        if template_type == "L":
            # Groups the FT_EN column by CAS numbers (i.e., chemical identifiers). 
            #Returns a dictionary where - keys = CAS numbers; values = lists of row indices in the DataFrame where that CAS occurs.
            grp_indcs = df["FT_EN"].groupby(df["CAS"]).groups
            
            # Filters g to only include a predefined list of CAS numbers from cas_numbers.
            # If a CAS number is not found in the original group, it assigns an empty string "".
            grp_indcs = {cas: grp_indcs.get(cas, "") for cas in cas_numbers}
            
        elif template_type == "K":
            grp_indcs = df["FT_EN"].groupby(df["CAS"]).groups 
            grp_indcs = {cas: grp_indcs.get(cas, "") for cas in cas_numbers}
        else:
            raise Exception("Template not found")    
        # Deduplicate English footnotes
        fts_combined_for_all_cas = {}
        for cas, indices in grp_indcs.items():
            # Retrieve the actual footnote texts from the DataFrame for each index, filtering out None values
            footnotes = []
            for i in indices:
                ft_en = df.loc[i, "FT_EN"]
                ft_hcodes = df.loc[i, "FT_HZ"]
                
                if ft_en is not None:
                    footnotes.append(ft_en)
                elif ft_hcodes is not None:
                    footnotes.append(ft_hcodes)
            
            # Deduplicate the texts while preserving their order
            deduped = list(dict.fromkeys(footnotes))
            # If deduped is nonempty, assign it; otherwise, assign an empty string
            fts_combined_for_all_cas[cas] = deduped if deduped else None
    
    # extend hazard codes with their description 
    for k, v in fts_combined_for_all_cas.items():
        fts_combined_for_all_cas[k] = extend_ft_hazard_code_if_any(df_hazard, v, language)
    
    for cas in cas_numbers:
        
        s = Substance()

        s.cas = str(cas).lstrip("0")
        s.canonical_smiles = df.loc[df["CAS"] == cas, "SMILES"].iloc[0] if not pd.isna(df.loc[df["CAS"] == cas, "SMILES"].iloc[0]) else ''
        s.iupac_name_de = df.loc[df["CAS"] == cas, "IUPAC_DE"].iloc[0] if not pd.isna(df.loc[df["CAS"] == cas, "IUPAC_DE"].iloc[0]) else ''
        s.iupac_name_en = df.loc[df["CAS"] == cas, "IUPAC_EN"].iloc[0] if not pd.isna(df.loc[df["CAS"] == cas, "IUPAC_EN"].iloc[0]) else ''

        s.fcm = df.loc[df["CAS"] == cas, "FCM"].iloc[0] if not pd.isna(df.loc[df["CAS"] == cas, "FCM"].iloc[0]) else ''
        s.cramer = df.loc[df["CAS"] == cas, "Cramer"].iloc[0] if not pd.isna(df.loc[df["CAS"] == cas, "Cramer"].iloc[0]) else ''
        
        if language == "DE":

            match_condition = (df_cl["CAS"] == cas) & ((df_cl["TypeFlag_CL"] == template_type) | (df_cl["TypeFlag_CL"] == "A"))
            filtered = df_cl.loc[match_condition, "CL_DE"]
            if not filtered.empty:
                if len(filtered) > 1:
                    print(f"Multiple Classifications found for CAS {s.cas}")
                    s.cl_de = "; ".join(filtered.dropna().astype(str))  # Combine multiple entries
                else:
                    s.cl_de = filtered.iloc[0] if not pd.isna(filtered.iloc[0]) else []
            else:
                s.cl_de = []

            
            # fts combined
            s.ft_de =(
                [item for item in fts_combined_for_all_cas[cas] if item is not None]
                if fts_combined_for_all_cas[cas]
                else []
            )
            
        elif language == "EN":
            # s.cl_de = df.loc[df["CAS"] == cas, "CL_EN"].iloc[0] if pd.isna(df.loc[df["CAS"] == cas, "CL_EN"].iloc[0]) else ''
            match_condition = (df_cl["CAS"] == cas) & ((df_cl["TypeFlag_CL"] == template_type) | (df_cl["TypeFlag_CL"] == "A"))
            filtered = df_cl.loc[match_condition, "CL_EN"]
            if not filtered.empty:
                if len(filtered) > 1:
                    # print(f"Multiple Classifications found for CAS {s.cas}")
                    s.cl_en = "; ".join(filtered.dropna().astype(str))  # Combine multiple entries
                else:
                    s.cl_en = filtered.iloc[0] if not pd.isna(filtered.iloc[0]) else []
            else:
                s.cl_en = []
                
            # fts combined
            # s.ft_de = "\n".join(fts_combined_for_all_cas[cas]) if "None" in fts_combined_for_all_cas[cas] else ''
            s.ft_en = (
                [item for item in fts_combined_for_all_cas[cas] if item is not None]
                if fts_combined_for_all_cas[cas]
                else []
            ) 
        db_substances[s.cas] = s
    print(db_substances)
    return df_eu, df_eu_grp
    
def extend_ft_hazard_code_if_any(df_hazard, ft, language):
    """
    Expands any H-code entries in a functional-type list with their full description.

    EU hazard codes such as ``'H302'`` are stored as bare codes in the database.
    This function detects them and appends the language-appropriate description so
    that the Word report shows ``'H302: Harmful if swallowed'`` rather than just
    ``'H302'``.

    Args:
        df_hazard (pd.DataFrame): The ``Hcodes`` sheet from the NIAS DB, with
            columns ``'Code'``, ``'Description_DE'``, ``'Description_EN'``.
        ft (str | list | None): A single functional-type string or a list of
            such strings.  ``None`` is handled gracefully.
        language (str): ``'DE'`` or ``'EN'``; selects the description column.

    Returns:
        str | list | None: The input with any H-codes replaced by
        ``'<code>: <description>'`` strings.  Returns ``None`` if ``ft`` is
        falsy.
    """
    # check if ft matches the Hcode regex
    if ft:
        if isinstance(ft, list) and len(ft) > 0:
            ft = [get_hazard_descp(df_hazard, language, f) for f in ft]
            return ft
        return get_hazard_descp(df_hazard, language, ft)
        # if yes then add the description


def get_hazard_descp(df_hazard, language, code):
    """
    Returns the full description for a single H-code, or the original code if it is not
    an H-code or is not found in the database.

    A string is treated as an H-code when it:
    - Starts with the letter ``'H'`` followed by alphanumeric or ``'+'`` characters.
    - Has a total length shorter than 12 characters (guards against false positives
      such as long INCI or classification strings).

    Args:
        df_hazard (pd.DataFrame): The ``Hcodes`` sheet from the NIAS DB.
        language (str): ``'DE'`` or ``'EN'``; selects the description column.
        code (str): The functional-type text to inspect.

    Returns:
        str: ``'<code>: <description>'`` if found; the original ``code`` otherwise.
             If the code is present in the DB but the lookup index raises an
             ``IndexError``, the suffix ``'$$$$$$$$$$$$Not found in database. Add manually'``
             is appended as a visual warning in the report.
    """
    if re.match(r"^H[a-zA-Z0-9+ ]*", code) and len(code) < 12 :
        try:
            if language == "DE":
                return code + ": " + df_hazard.loc[df_hazard["Code"] == code, "Description_DE"].iloc[0]
            elif language == "EN":
                return code + ": " + df_hazard.loc[df_hazard["Code"] == code, "Description_EN"].iloc[0]
        except IndexError as e:
            print(e)
            return code + ": $$$$$$$$$$$$Not found in database. Add manually" 
    return code

def parse_table(leading_cas_number, t, cols, ft_text, lab_substances):
    """
    Parses a specific table in a Word document to extract substance information.

    Args:
        leading_cas_number : Callable to match regex
        t (docx.Table): The table object to parse.
        cols (dict): Mapping of column indices to data attributes.
            Example: {0: 'Name', 1: 'CAS', 2: 'FCM'}
        ft_text (dict): Dictionary of footnote texts indexed by their identifiers.
        lab_substances (dict): Dictionary of substances from the corresponding Excel file.

    Returns:
        dict: Parsed substances, indexed by CAS numbers.
    """
    
    substances = {}
    for row in range(1, len(t.rows)):
        #  # uninitialized Substance object S
        s = Substance()
        name = ''
        fcm = None
        ft_list = []
        cls_list = []
        inci = ''

        for col in cols.keys():
            col_def = cols[col]
            (text, ft) = get_text_and_superscript(t.cell(row, col))
            text = text.strip()
            fts = [x.strip() for x in ft.split(',')]

            if col_def == 'Name' or col_def == 'Name_CAS':
                name = text
                for ft_char in fts:
                    if len(ft_char) > 0:
                        ft_list.append(ft_text[ft_char.strip()])
                if col_def == 'Name_CAS':
                    m = re.search('\\(CAS (.*)\\)', t.cell(row, col).text)
                    if m is not None:
                        cas = m.group(1)
                        cas = re.sub(r'^(\d+)', leading_cas_number, cas)
                    else:
                        cas = '-'
            elif col_def == 'CAS':
                cas = text
                if cas != '-':
                    cas = re.sub(r'^(\d+)', leading_cas_number, cas)
            elif col_def == 'FCM':
                fcm = text
            elif col_def == 'CLS':
                (cls_list, inci) = get_cl_and_inci(t.cell(row, col).text)
        
        
        # ----- TODO: At this point the docx and xlsx files are coupled
        #  This is why searching is done in pairs - here the substances from DB.xlsx are assigned the cas, fcm, ft, cl and inci from doc file (if found)
        if cas == '-':
            rt_col = [i for i in cols if cols[i] == 'RT']
            if len(rt_col) > 0:
                rt_col = rt_col[0]

                rts = t.cell(row, rt_col).text.strip()
                rts = rts.split('\n')
                rts = [float(x.replace(',', '.')) for x in rts]
                cas_list = []

                
                for rt in rts:
                    for rt_lab in lab_substances.keys():
                        if abs(rt - rt_lab) < 0.01:
                            s = lab_substances[rt_lab]
                            cas_leading_zeroes = re.sub(r'^(\d+)', leading_cas_number, s.cas)
                            s.cas = cas_leading_zeroes
                            s.fcm = fcm
                            s.ft = ft_list
                            s.cl = []
                            s.inci = ''
                            print(f' {s.name}\t{s.cas}\t{s.fcm}\t{s.ft}\t{s.cl}\t{s.inci}')
                            substances[s.cas] = s
                            break
        else:
            s = Substance()
            s.name = name
            s.cas = cas
            s.fcm = fcm
            s.ft = ft_list
            s.cl = cls_list
            s.inci = inci
            print(f' {s.name}\t{s.cas}\t{s.fcm}\t{s.ft}\t{s.cl}\t{s.inci}')
            substances[s.cas] = s

    return substances


