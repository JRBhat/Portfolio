"""
AutoFotoValidation — GCP-compliant image folder validation utility.

Pipeline:
  1. Create / locate the Process_Validation_file.xlsx workbook.
  2. Log source folder paths into the workbook.
  3. Copy source folder contents to a $Basic_cleaned staging folder via TeraCopy.
  4. Rename folders (add/remove the '$' active-prefix sentinel) to track state.
  5. Repeat for each user-created destination folder until the study is complete.
"""

import os
import sys
import subprocess
import datetime
import shutil
import glob
from pathlib import Path
from openpyxl import Workbook, load_workbook, utils
from openpyxl.styles import Font

# ── Constants ────────────────────────────────────────────────────────────────
ACTIVE_PREFIX = "$"
DEFAULT_STUDY_FOLDER = r"D:\test_software"
WORKBOOK_NAME = "Process_Validation_file.xlsx"
WORKBOOK_SHEET = "Main_sheet"
TEMP_SOURCE_LIST_NAME = "temp_source_for_tercpy.txt"
LOG_FOLDER_NAME = "logs"
BASIC_CLEANED_FOLDER = "$Basic_cleaned"
TERACOPY_EXE = r"C:\Program Files\TeraCopy\TeraCopy.exe"
MAX_LOG_ROWS = 100
COL_FOLDER_NAME = "A"
COL_SOURCE_LINK = "B"
COL_DEST_LINK = "C"
COL_MTIME = "D"
COL_REPORT_LINK = "E"
COLOR_ERROR = "FFFF0000"
COLOR_OK = "FF008000"
SKIP_SUFFIXES = (".xlsx", ".txt")
MSG_NA_AT_START = "Not Applicable during Start"
MSG_NA = "Not applicable"
MSG_PROTOCOL_PLACEHOLDER = "backup_server download protocol goes here"
MSG_PROTOCOL_ERROR = "Error. Protocol not created. "
MSG_FOLDER_REPEATED = "Erroneous folder deleted and process repeated"

# ── Helpers ──────────────────────────────────────────────────────────────────
def unmark(path: Path) -> Path:
    return path.with_name(path.name.replace(ACTIVE_PREFIX, ""))

def mark(path: Path) -> Path:
    return path.with_name(ACTIVE_PREFIX + path.name)


def main():
    # reads the main study path ; prerequisite: study path must already contain the folder with backup_server downloaded images
    main_study_folder_path = DEFAULT_STUDY_FOLDER
    common_tcpy_paths = [os.path.join(os.environ.get("APPDATA", ""), "TeraCopy", "Reports")]
    none_counter = 0

    ignore_list = []
    tmp_file_list = []

    preliminary_cleaning_done = input("Continue where you left off? (y/n): ").strip().lower() == "y"
    if preliminary_cleaning_done:
        for entry in os.listdir(main_study_folder_path):
            if ACTIVE_PREFIX not in entry and not entry.endswith(SKIP_SUFFIXES):
                ignore_list.append(entry)

    # creates a log folder to store teracopy reports
    log_folder_path = os.path.join(main_study_folder_path, LOG_FOLDER_NAME)
    if os.path.isdir(log_folder_path):
        print("folder_already exists, skipping step")
    else:
        os.mkdir(log_folder_path)
    ignore_list.append(LOG_FOLDER_NAME)

    # creates an excel file and logs the folder name and it's pathname
    excel_path = os.path.join(main_study_folder_path, WORKBOOK_NAME)
    if os.path.isfile(excel_path):
        print("file already exists, skipping step")
        none_counter = 2
    else:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = WORKBOOK_SHEET
        workbook.save(excel_path)

        for file in os.listdir(main_study_folder_path):
            if not file.endswith(SKIP_SUFFIXES):
                none_counter = log_paths_to_excel(excel_path, file, os.path.join(main_study_folder_path, file), None, None, none_counter, log_folder_path)

    # creates a copy of the folder already present
        # copies all images from existing madasa downloaded folder

    while True:
        tmp_file_list = []
        if not preliminary_cleaning_done:
            for file in os.listdir(main_study_folder_path):
                if not file.endswith(SKIP_SUFFIXES) and file not in ignore_list:

                    orig_path = os.path.join(main_study_folder_path, file)
                    # copying first file to basic cleaned folder using teracopy
                    basic_cleaned_folder_name = BASIC_CLEANED_FOLDER
                    basic_cleaned_folder_path = os.path.join(main_study_folder_path, basic_cleaned_folder_name)
                    os.mkdir(basic_cleaned_folder_path)

                    path_to_copy_report = copy_folder_contents_using_teracopy(orig_path, basic_cleaned_folder_path, main_study_folder_path, tmp_file_list, common_tcpy_paths, log_folder_path)
                    # cleaning .ini and .xml files if any

                    # logging in excel file
                    none_counter = log_paths_to_excel(excel_path, basic_cleaned_folder_name.replace(ACTIVE_PREFIX, ""), orig_path, basic_cleaned_folder_path, path_to_copy_report, none_counter, log_folder_path)

                    # orignal folder added to ignore list
                    ignore_list.append(file)

                    # flag updated indicates end of preliminary process
                    preliminary_cleaning_done = True
                    break

        elif preliminary_cleaning_done:  # for user created directories
            while True:
                new_folder_creation_prompt = input("Have you created a new folder?(y/n) ")
                if new_folder_creation_prompt.lower() == "y":
                    for file in os.listdir(main_study_folder_path):
                        if not file.endswith(SKIP_SUFFIXES) and file not in ignore_list:
                            if ACTIVE_PREFIX in file:
                                filename_to_be_copied = file
                                ignore_list.append(file)
                                continue
                            copy_from = os.path.join(main_study_folder_path, filename_to_be_copied)
                            copy_to = os.path.join(main_study_folder_path, file)

                            path_to_copy_report = copy_folder_contents_using_teracopy(copy_from, copy_to, main_study_folder_path, tmp_file_list, common_tcpy_paths, log_folder_path)

                            none_counter = log_paths_to_excel(excel_path, os.path.basename(copy_to), copy_from, copy_to, path_to_copy_report, none_counter, log_folder_path)

                            Path(copy_from).rename(unmark(Path(copy_from)))
                            Path(copy_to).rename(mark(Path(copy_to)))
                            ignore_list = [item.replace(ACTIVE_PREFIX, "") if ACTIVE_PREFIX in item else item for item in ignore_list]
                            break

                elif new_folder_creation_prompt.lower() == "n":
                    while True:
                        check_file_ignore = input("Do you want to remove a specific folder from the ignore list?(y/n)")
                        if check_file_ignore.lower() == "y":
                            print("Please delete the folder manually and assign $ to previous source folder")
                            input("Press enter after performing all corrective actions")
                            break
                        elif check_file_ignore.lower() == "n":
                            for file in os.listdir(main_study_folder_path):
                                if ACTIVE_PREFIX in file:
                                    clean_filename = file.replace(ACTIVE_PREFIX, "")
                                    os.rename(os.path.join(main_study_folder_path, file), os.path.join(main_study_folder_path, clean_filename))
                            print("No more tasks created, process ending.")
                            sys.exit(0)


def log_paths_to_excel(excel_path, folder_name, copyfrom_fpathname, copyto_fpathname, report_path, none_counter, folder_path_to_log):
    """
    Log folder copy details into the Process Validation workbook.

    Args:
        excel_path: Path to the Excel workbook file.
        folder_name: Display name for the folder (written to column A).
        copyfrom_fpathname: Source folder path (hyperlinked in column B).
        copyto_fpathname: Destination folder path, or None for initial entries.
        report_path: TeraCopy report filename, or None if unavailable.
        none_counter: Tracks how many None-report entries have been written.
        folder_path_to_log: Directory where TeraCopy reports are stored.

    Returns:
        Updated none_counter value.
    """
    workbook = load_workbook(excel_path)
    worksheet = workbook[WORKBOOK_SHEET]

    if copyto_fpathname is None:
        for row in worksheet.iter_cols(min_row=1, min_col=2, max_row=MAX_LOG_ROWS, max_col=2):
            for row_index, cell in enumerate(row, start=1):
                if cell.value is not None:
                    continue
                else:
                    letter = utils.get_column_letter(cell.column)
                    print(f"{letter}{row_index}")
                    cell.hyperlink = str(copyfrom_fpathname)
                    cell.style = "Hyperlink"
                    worksheet[f'{COL_FOLDER_NAME}{row_index}'].value = folder_name.replace(ACTIVE_PREFIX, "")
                    worksheet[f'{COL_DEST_LINK}{row_index}'].value = MSG_NA_AT_START
                    worksheet[f'{COL_MTIME}{row_index}'].value = datetime.datetime.fromtimestamp(os.path.getmtime(copyfrom_fpathname))
                    if report_path is not None:
                        worksheet[f'{COL_REPORT_LINK}{row_index}'].hyperlink = str(os.path.join(folder_path_to_log, report_path))
                        worksheet[f'{COL_REPORT_LINK}{row_index}'].style = "Hyperlink"
                    elif report_path is None and none_counter == 0:
                        worksheet[f'{COL_REPORT_LINK}{row_index}'].value = MSG_NA
                        none_counter += 1
                    elif report_path is None and none_counter == 1:
                        worksheet[f'{COL_REPORT_LINK}{row_index}'].value = MSG_PROTOCOL_PLACEHOLDER
                        none_counter += 1
                    workbook.save(excel_path)
                    break
            break
        return none_counter

    else:
        for row in worksheet.iter_cols(min_row=1, min_col=2, max_row=MAX_LOG_ROWS, max_col=2):
            for row_index, cell in enumerate(row, start=1):
                if cell.value is not None:
                    continue
                else:
                    letter = utils.get_column_letter(cell.column)
                    print(f"{letter}{row_index}")
                    cell.hyperlink = str(copyfrom_fpathname).replace(ACTIVE_PREFIX, "")
                    cell.style = "Hyperlink"
                    worksheet[f'{COL_FOLDER_NAME}{row_index}'].value = folder_name.replace(ACTIVE_PREFIX, "")
                    worksheet[f'{COL_DEST_LINK}{row_index}'].hyperlink = copyto_fpathname.replace(ACTIVE_PREFIX, "")
                    worksheet[f'{COL_DEST_LINK}{row_index}'].style = "Hyperlink"
                    worksheet[f'{COL_MTIME}{row_index}'].value = datetime.datetime.fromtimestamp(os.path.getmtime(copyfrom_fpathname))
                    if report_path is not None:
                        worksheet[f'{COL_REPORT_LINK}{row_index}'].hyperlink = str(os.path.join(folder_path_to_log, report_path))
                        worksheet[f'{COL_REPORT_LINK}{row_index}'].style = "Hyperlink"
                    elif report_path is None and none_counter == 0:
                        worksheet[f'{COL_REPORT_LINK}{row_index}'].value = MSG_NA
                        none_counter += 1
                    elif report_path is None and none_counter == 1:
                        worksheet[f'{COL_REPORT_LINK}{row_index}'].value = MSG_PROTOCOL_PLACEHOLDER
                        none_counter += 1
                    elif report_path is None and none_counter > 1:
                        worksheet[f'{COL_REPORT_LINK}{row_index}'].value = MSG_PROTOCOL_ERROR
                        worksheet[f'{COL_REPORT_LINK}{row_index}'].font = Font(color=COLOR_ERROR)
                        worksheet[f'{COL_DEST_LINK}{row_index}'].hyperlink = None
                        worksheet[f'{COL_DEST_LINK}{row_index}'].value = MSG_FOLDER_REPEATED
                        worksheet[f'{COL_DEST_LINK}{row_index}'].font = Font(color=COLOR_OK)
                        none_counter += 1
                    workbook.save(excel_path)
                    break
            break
        return none_counter


def copy_folder_contents_using_teracopy(from_path, to_path, main_path, temp_list_with_paths, common_std_tcpy_paths, destin_path_to_move_report):
    """
    Copy all files from from_path into to_path using TeraCopy, then move the TeraCopy
    report to the log folder.

    Args:
        from_path: Source directory whose contents are copied.
        to_path: Destination directory.
        main_path: Study root directory (used to write the temp file list).
        temp_list_with_paths: Accumulator list for file paths (cleared internally).
        common_std_tcpy_paths: List of directories where TeraCopy stores its reports.
        destin_path_to_move_report: Directory where the report is moved after copying.

    Returns:
        Filename of the TeraCopy report on success, or None on failure.
    """
    expected_files = set()
    temp_list_with_paths = []
    for filename in os.listdir(from_path):
        file_to_copy = os.path.join(from_path, filename)
        path_to_txt_file = os.path.join(main_path, TEMP_SOURCE_LIST_NAME)
        temp_list_with_paths.append(file_to_copy)
        expected_files.add(file_to_copy)
    with open(path_to_txt_file, 'a') as txt_handle:
        for path in temp_list_with_paths:
            txt_handle.write(path + "\n")
    subprocess.run([TERACOPY_EXE, "Copy", f"*{path_to_txt_file}", to_path, "/NoClose"], check=True)
    log_file_path = get_copy_protocol_from_stdpath(common_std_tcpy_paths, expected_files, destin_path_to_move_report)
    delete_txt_files_in_dir(main_path)
    if log_file_path != 0:
        return log_file_path
    else:
        print("log file not found. Unsuccessful copying")


def delete_txt_files_in_dir(path):
    """
    Delete all .txt files in path.

    Args:
        path: Directory to clean.
    """
    for file in os.listdir(path):
        if file.endswith(".txt"):
            os.remove(os.path.join(path, file))


def get_copy_protocol_from_stdpath(common_teracopy_protocol_path_list, expected_files, destination_path):
    """
    Find the most recent TeraCopy CSV report whose file list matches expected_files and
    move it to destination_path.

    Args:
        common_teracopy_protocol_path_list: Directories to search for TeraCopy reports.
        expected_files: Set of file paths that must appear in the report to be a match.
        destination_path: Directory to move the matched report into.

    Returns:
        Filename of the matched report, or 0 if no match was found.
    """
    # assign std_path where teracopy stores the pdf report, to a variable called tcpy_std_path
    for each_path in common_teracopy_protocol_path_list:
        try:
            reported_files = set()
            each_path = os.path.join(os.environ.get("APPDATA", ""), "TeraCopy", "Reports")
            csv_files = glob.glob(os.path.join(each_path, "*.csv"))
            max_file_csv = max(csv_files, key=os.path.getmtime)
            with open(max_file_csv, "r") as txt_handle:
                lines = txt_handle.readlines()
                for line in lines:
                    for path in line.split(","):
                        if "\\" in path:
                            print(path)
                            reported_files.add(path.replace("\n", ""))
            if expected_files == reported_files:
                shutil.move(max_file_csv, destination_path)
                print("Files moved as a result of successful 1st loop")
                break
        except Exception:
            print("File not found using glob, switching to manual iteration...")
            for file in os.listdir(each_path):
                with open(os.path.join(each_path, file), "r") as txt_handle:
                    lines = txt_handle.readlines()
                    for line in lines:
                        for path in line.split(","):
                            if "\\" in path:
                                print(path)
                                reported_files.add(path.replace("\n", ""))
            if expected_files == reported_files:
                shutil.move(os.path.join(each_path, file), destination_path)
                print("Files moved as a result of successful 2nd loop")
                break
        else:
            print("Operation not successful")
    try:
        if max_file_csv:
            return os.path.basename(max_file_csv)
        elif file:
            return file
    except UnboundLocalError:
        print("Report not created")
        return 0


if __name__ == "__main__":
    main()

# teracopy reports save location --> send to new log folder in main_study folder
# Default location: %APPDATA%\TeraCopy\Reports
