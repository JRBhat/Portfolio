"""
AutoFotoValidation — basic image folder validation (without TeraCopy reporting).

Pipeline:
  1. Create the Process_Validation_file.xlsx workbook.
  2. Log source folder paths into the workbook.
  3. Copy source folder contents to a $Basic_cleaned staging folder via TeraCopy.
  4. Rename folders (add/remove the '$' active-prefix sentinel) to track state.
  5. Repeat for each user-created destination folder until the study is complete.
"""

import os
import sys
import subprocess
import datetime
from pathlib import Path
from openpyxl import Workbook, load_workbook, utils

# ── Constants ────────────────────────────────────────────────────────────────
ACTIVE_PREFIX = "$"
DEFAULT_STUDY_FOLDER = r"D:\test_software"
WORKBOOK_NAME = "Process_Validation_file.xlsx"
WORKBOOK_SHEET = "Main_sheet"
TEMP_SOURCE_LIST_NAME = "temp_source_for_tercpy.txt"
BASIC_CLEANED_FOLDER = "$Basic_cleaned"
TERACOPY_EXE = r"C:\Program Files\TeraCopy\TeraCopy.exe"
MAX_LOG_ROWS = 100
COL_FOLDER_NAME = "A"
COL_DEST_LINK = "C"
COL_MTIME = "D"
SKIP_SUFFIXES = (".xlsx", ".txt")
MSG_NA_AT_START = "Not Applicable during Start"


# ── Helpers ──────────────────────────────────────────────────────────────────
def unmark(path: Path) -> Path:
    return path.with_name(path.name.replace(ACTIVE_PREFIX, ""))

def mark(path: Path) -> Path:
    return path.with_name(ACTIVE_PREFIX + path.name)


def main():
    # reads the main study path ; prerequisite: study path must already contain the folder with madase downloaded images
    main_study_folder_path = DEFAULT_STUDY_FOLDER
    common_tcpy_paths = []
    # 0 = fresh run, 1 = resume
    preliminary_cleaning_done = 0

    # creates an excel file and logs the folder name and it's pathname
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = WORKBOOK_SHEET
    excel_path = os.path.join(main_study_folder_path, WORKBOOK_NAME)
    workbook.save(excel_path)
    ignore_list = []
    tmp_file_list = []
    for file in os.listdir(main_study_folder_path):
        if not file.endswith(SKIP_SUFFIXES):
            log_paths_to_excel(excel_path, file, os.path.join(main_study_folder_path, file), None)

    # creates a copy of the folder already present
        # copies all images from existing madasa downloaded folder

    while True:
        tmp_file_list = []
        if preliminary_cleaning_done == 0:
            for file in os.listdir(main_study_folder_path):
                if not file.endswith(SKIP_SUFFIXES) and file not in ignore_list:

                    orig_path = os.path.join(main_study_folder_path, file)
                    # copying first file to basic cleaned folder using teracopy
                    basic_cleaned_folder_name = BASIC_CLEANED_FOLDER
                    basic_cleaned_folder_path = os.path.join(main_study_folder_path, basic_cleaned_folder_name)
                    os.mkdir(basic_cleaned_folder_path)

                    copy_folder_contents_using_teracopy(orig_path, basic_cleaned_folder_path, main_study_folder_path, tmp_file_list)
                    # cleaning .ini and .xml files if any

                    # logging in excel file
                    log_paths_to_excel(excel_path, basic_cleaned_folder_name.replace(ACTIVE_PREFIX, ""), orig_path, basic_cleaned_folder_path)

                    # orignal folder added to ignore list
                    ignore_list.append(file)

                    # flag updated indicates end of preliminary process
                    preliminary_cleaning_done = 1
                    break

        elif preliminary_cleaning_done == 1:  # for user created directories
            new_folder_creation_prompt = input("Have you created a new folder?(y/n) ")
            if new_folder_creation_prompt.lower() == "y":
                for file in os.listdir(main_study_folder_path):
                    if not file.endswith(SKIP_SUFFIXES) and file not in ignore_list:
                        if ACTIVE_PREFIX in file:
                            filename_to_be_copied = file
                            ignore_list.append(file)
                            continue
                        copy_from = os.path.join(main_study_folder_path, filename_to_be_copied)
                        copy_to = os.path.join(main_study_folder_path, file)

                        copy_folder_contents_using_teracopy(copy_from, copy_to, main_study_folder_path, tmp_file_list)

                        log_paths_to_excel(excel_path, os.path.basename(copy_to), copy_from, copy_to)
                        Path(copy_from).rename(unmark(Path(copy_from)))
                        Path(copy_to).rename(mark(Path(copy_to)))
                        ignore_list = [item.replace(ACTIVE_PREFIX, "") if ACTIVE_PREFIX in item else item for item in ignore_list]
                        break

            elif new_folder_creation_prompt.lower() == "n":
                for file in os.listdir(main_study_folder_path):
                    if ACTIVE_PREFIX in file:
                        clean_filename = file.replace(ACTIVE_PREFIX, "")
                        os.rename(os.path.join(main_study_folder_path, file), os.path.join(main_study_folder_path, clean_filename))
                print("No more tasks created, process ending.")
                sys.exit(0)


def log_paths_to_excel(excel_path, folder_name, copyfrom_fpathname, copyto_fpathname):
    """
    Log folder copy details into the Process Validation workbook.

    Args:
        excel_path: Path to the Excel workbook file.
        folder_name: Display name for the folder (written to column A).
        copyfrom_fpathname: Source folder path (hyperlinked in column B).
        copyto_fpathname: Destination folder path, or None for initial entries.
    """
    workbook = load_workbook(excel_path)
    worksheet = workbook[WORKBOOK_SHEET]

    if copyto_fpathname is None:
        for row in worksheet.iter_cols(min_row=1, min_col=2, max_row=MAX_LOG_ROWS, max_col=2):
            for row_index, cell in enumerate(row, start=1):
                if cell.value is not None:
                    continue
                else:
                    letter = utils.get_column_letter(cell.column)
                    print(f"{letter}{row_index}")
                    cell.hyperlink = str(copyfrom_fpathname)
                    worksheet[f'{COL_FOLDER_NAME}{row_index}'].value = folder_name.replace(ACTIVE_PREFIX, "")
                    worksheet[f'{COL_DEST_LINK}{row_index}'].value = MSG_NA_AT_START
                    worksheet[f'{COL_MTIME}{row_index}'].value = datetime.datetime.fromtimestamp(os.path.getmtime(copyfrom_fpathname))
                    workbook.save(excel_path)
                    break
            break

    else:
        for row in worksheet.iter_cols(min_row=1, min_col=2, max_row=MAX_LOG_ROWS, max_col=2):
            for row_index, cell in enumerate(row, start=1):
                if cell.value is not None:
                    continue
                else:
                    letter = utils.get_column_letter(cell.column)
                    print(f"{letter}{row_index}")
                    cell.hyperlink = str(copyfrom_fpathname).replace(ACTIVE_PREFIX, "")
                    worksheet[f'{COL_FOLDER_NAME}{row_index}'].value = folder_name.replace(ACTIVE_PREFIX, "")
                    worksheet[f'{COL_DEST_LINK}{row_index}'].hyperlink = copyto_fpathname.replace(ACTIVE_PREFIX, "")
                    worksheet[f'{COL_MTIME}{row_index}'].value = datetime.datetime.fromtimestamp(os.path.getmtime(copyfrom_fpathname))
                    workbook.save(excel_path)
                    break
            break


def copy_folder_contents_using_teracopy(from_path, to_path, main_path, temp_list_with_paths):
    """
    Copy all files from from_path into to_path using TeraCopy.

    Args:
        from_path: Source directory whose contents are copied.
        to_path: Destination directory.
        main_path: Study root directory (used to write the temp file list).
        temp_list_with_paths: Accumulator list for file paths (cleared internally).
    """
    for filename in os.listdir(from_path):
        file_to_copy = os.path.join(from_path, filename)
        path_to_txt_file = os.path.join(main_path, TEMP_SOURCE_LIST_NAME)
        temp_list_with_paths.append(file_to_copy)
    with open(path_to_txt_file, 'a') as txt_handle:
        for path in temp_list_with_paths:
            txt_handle.write(path + "\n")
    subprocess.run([TERACOPY_EXE, "Copy", f"*{path_to_txt_file}", to_path, "/NoClose"], check=True)
    delete_txt_files_in_dir(main_path)


def delete_txt_files_in_dir(path):
    """
    Delete all .txt files in path.

    Args:
        path: Directory to clean.
    """
    for file in os.listdir(path):
        if file.endswith(".txt"):
            os.remove(os.path.join(path, file))


if __name__ == "__main__":
    main()
